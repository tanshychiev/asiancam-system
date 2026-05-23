from decimal import Decimal, InvalidOperation
from datetime import date, datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.models import Company

from .forms import (
    ChartOfAccountForm,
    JournalEntryForm,
    JournalEntryLineFormSet,
)
from .models import (
    AccountCustomer,
    AccountItem,
    BulkImportLog,
    ChartOfAccount,
    JournalEntry,
    JournalEntryLine,
)
from .sample_coa import create_sample_coa


# =========================================================
# HELPERS
# =========================================================

def get_selected_company(request):
    company_id = request.session.get("selected_company_id")

    if not company_id:
        return None

    return Company.objects.filter(
        id=company_id,
        is_active=True,
    ).first()


def can_access_company(user, company):
    if not company:
        return False

    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)

    if profile and getattr(profile, "user_type", None) == "client":
        return profile.company_id == company.id

    if hasattr(company, "assigned_staff"):
        return company.assigned_staff.filter(id=user.id).exists()

    return False


def require_company_access(request):
    company = get_selected_company(request)

    if not company:
        messages.warning(request, "Please select a company first.")
        return None, redirect("company_list")

    if not can_access_company(request.user, company):
        messages.error(request, "You do not have permission to access this company.")
        return None, redirect("company_list")

    return company, None


def to_decimal(value):
    if value in [None, ""]:
        return Decimal("0.00")

    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0.00")


def parse_excel_date(value):
    if not value:
        return timezone.localdate()

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError:
        return timezone.localdate()


def get_posted_status():
    return getattr(JournalEntry, "STATUS_POSTED", "posted")


def get_draft_status():
    return getattr(JournalEntry, "STATUS_DRAFT", "draft")


# =========================================================
# CHART OF ACCOUNTS
# =========================================================

@login_required
def chart_of_accounts(request):
    company, response = require_company_access(request)
    if response:
        return response

    query = (request.GET.get("q") or "").strip()
    account_type = (request.GET.get("account_type") or "").strip()
    report_type = (request.GET.get("report_type") or "").strip()

    accounts = ChartOfAccount.objects.filter(company=company)

    if query:
        accounts = accounts.filter(
            Q(code__icontains=query)
            | Q(name__icontains=query)
            | Q(local_name__icontains=query)
            | Q(description__icontains=query)
        )

    if account_type:
        accounts = accounts.filter(account_type=account_type)

    if report_type:
        accounts = accounts.filter(report_type=report_type)

    accounts = accounts.select_related("parent").order_by("code")

    total_accounts = accounts.count()
    active_accounts = accounts.filter(is_active=True).count()
    group_accounts = accounts.filter(is_group=True).count()

    account_type_stats = (
        ChartOfAccount.objects
        .filter(company=company, is_active=True)
        .values("account_type")
        .annotate(total=Count("id"))
        .order_by("account_type")
    )

    has_accounts = ChartOfAccount.objects.filter(company=company).exists()

    return render(request, "accounting/chart_of_accounts.html", {
        "company": company,
        "accounts": accounts,
        "query": query,
        "account_type": account_type,
        "report_type": report_type,
        "total_accounts": total_accounts,
        "active_accounts": active_accounts,
        "group_accounts": group_accounts,
        "account_type_stats": account_type_stats,
        "has_accounts": has_accounts,
        "account_type_choices": ChartOfAccount.ACCOUNT_TYPE_CHOICES,
        "report_type_choices": ChartOfAccount.REPORT_TYPE_CHOICES,
    })


@login_required
def setup_sample_coa(request):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method == "POST":
        created_count = create_sample_coa(company)

        if created_count > 0:
            messages.success(request, f"Sample Chart of Accounts created: {created_count} accounts.")
        else:
            messages.warning(request, "Sample Chart of Accounts already exists for this company.")

        return redirect("chart_of_accounts")

    return render(request, "accounting/setup_sample_coa.html", {
        "company": company,
    })


@login_required
def account_create(request):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method == "POST":
        form = ChartOfAccountForm(request.POST, company=company)

        if form.is_valid():
            account = form.save(commit=False)
            account.company = company
            account.save()

            messages.success(request, f"Account {account.code} - {account.name} created successfully.")
            return redirect("chart_of_accounts")
    else:
        form = ChartOfAccountForm(
            company=company,
            initial={
                "is_active": True,
                "normal_balance": "debit",
            },
        )

    return render(request, "accounting/account_form.html", {
        "form": form,
        "company": company,
        "page_title": "Create Account",
        "button_text": "Create Account",
    })


@login_required
def account_edit(request, account_id):
    company, response = require_company_access(request)
    if response:
        return response

    account = get_object_or_404(
        ChartOfAccount,
        id=account_id,
        company=company,
    )

    if request.method == "POST":
        form = ChartOfAccountForm(
            request.POST,
            instance=account,
            company=company,
        )

        if form.is_valid():
            form.save()
            messages.success(request, f"Account {account.code} updated successfully.")
            return redirect("chart_of_accounts")
    else:
        form = ChartOfAccountForm(
            instance=account,
            company=company,
        )

    return render(request, "accounting/account_form.html", {
        "form": form,
        "company": company,
        "account": account,
        "page_title": "Edit Account",
        "button_text": "Save Changes",
    })


@login_required
def account_toggle_active(request, account_id):
    company, response = require_company_access(request)
    if response:
        return response

    account = get_object_or_404(
        ChartOfAccount,
        id=account_id,
        company=company,
    )

    account.is_active = not account.is_active
    account.save(update_fields=["is_active"])

    if account.is_active:
        messages.success(request, f"{account.code} is now active.")
    else:
        messages.warning(request, f"{account.code} is now inactive.")

    return redirect("chart_of_accounts")


# =========================================================
# JOURNAL / ACCOUNTING DATA
# =========================================================

@login_required
def journal_list(request):
    company, response = require_company_access(request)
    if response:
        return response

    query = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    entries = JournalEntry.objects.filter(company=company)

    if query:
        entries = entries.filter(
            Q(reference_no__icontains=query)
            | Q(description__icontains=query)
            | Q(lines__account__name__icontains=query)
            | Q(lines__account__code__icontains=query)
        )

    if status:
        entries = entries.filter(status=status)

    if date_from:
        entries = entries.filter(entry_date__gte=date_from)

    if date_to:
        entries = entries.filter(entry_date__lte=date_to)

    entries = (
        entries
        .distinct()
        .prefetch_related("lines", "lines__account")
        .order_by("-entry_date", "-id")
    )

    total_entries = entries.count()
    posted_entries = entries.filter(status=get_posted_status()).count()
    draft_entries = entries.filter(status=get_draft_status()).count()

    return render(request, "accounting/journal_list.html", {
        "company": company,
        "entries": entries,
        "query": query,
        "status": status,
        "date_from": date_from,
        "date_to": date_to,
        "total_entries": total_entries,
        "posted_entries": posted_entries,
        "draft_entries": draft_entries,
        "status_choices": JournalEntry.STATUS_CHOICES,
    })


@login_required
def journal_create(request):
    company, response = require_company_access(request)
    if response:
        return response

    has_accounts = ChartOfAccount.objects.filter(
        company=company,
        is_active=True,
        is_group=False,
    ).exists()

    if not has_accounts:
        messages.warning(request, "Please create Chart of Accounts first.")
        return redirect("chart_of_accounts")

    journal_entry = JournalEntry(company=company, created_by=request.user)

    if request.method == "POST":
        form = JournalEntryForm(request.POST, instance=journal_entry)
        formset = JournalEntryLineFormSet(
            request.POST,
            instance=journal_entry,
            form_kwargs={"company": company},
        )

        if form.is_valid() and formset.is_valid():
            total_debit = Decimal("0.00")
            total_credit = Decimal("0.00")
            valid_lines = 0

            for line_form in formset:
                cleaned = getattr(line_form, "cleaned_data", None)

                if not cleaned:
                    continue

                if cleaned.get("DELETE", False):
                    continue

                account = cleaned.get("account")
                debit = cleaned.get("debit") or Decimal("0.00")
                credit = cleaned.get("credit") or Decimal("0.00")

                if account and (debit > 0 or credit > 0):
                    total_debit += debit
                    total_credit += credit
                    valid_lines += 1

                if debit > 0 and credit > 0:
                    messages.error(request, "One line cannot have both debit and credit.")
                    return redirect("journal_create")

            if valid_lines < 2:
                messages.error(request, "Journal entry must have at least 2 lines.")
            elif total_debit != total_credit:
                messages.error(
                    request,
                    f"Debit and Credit must be equal. Debit: {total_debit}, Credit: {total_credit}.",
                )
            elif total_debit <= 0:
                messages.error(request, "Total debit and credit must be more than zero.")
            else:
                with transaction.atomic():
                    entry = form.save(commit=False)
                    entry.company = company
                    entry.created_by = request.user
                    entry.save()

                    formset.instance = entry
                    formset.save()

                messages.success(request, f"Journal Entry {entry.entry_no} created successfully.")
                return redirect("journal_detail", entry_id=entry.id)

    else:
        form = JournalEntryForm(instance=journal_entry)
        formset = JournalEntryLineFormSet(
            instance=journal_entry,
            form_kwargs={"company": company},
        )

    return render(request, "accounting/journal_form.html", {
        "company": company,
        "form": form,
        "formset": formset,
        "page_title": "Create Journal Entry",
        "button_text": "Save Journal Entry",
    })


@login_required
def journal_detail(request, entry_id):
    company, response = require_company_access(request)
    if response:
        return response

    entry = get_object_or_404(
        JournalEntry.objects.prefetch_related("lines", "lines__account"),
        id=entry_id,
        company=company,
    )

    return render(request, "accounting/journal_detail.html", {
        "company": company,
        "entry": entry,
    })


@login_required
def journal_edit(request, entry_id):
    company, response = require_company_access(request)
    if response:
        return response

    entry = get_object_or_404(
        JournalEntry,
        id=entry_id,
        company=company,
    )

    if request.method == "POST":
        form = JournalEntryForm(request.POST, instance=entry)
        formset = JournalEntryLineFormSet(
            request.POST,
            instance=entry,
            form_kwargs={"company": company},
        )

        if form.is_valid() and formset.is_valid():
            total_debit = Decimal("0.00")
            total_credit = Decimal("0.00")
            valid_lines = 0

            for line_form in formset:
                cleaned = getattr(line_form, "cleaned_data", None)

                if not cleaned:
                    continue

                if cleaned.get("DELETE", False):
                    continue

                account = cleaned.get("account")
                debit = cleaned.get("debit") or Decimal("0.00")
                credit = cleaned.get("credit") or Decimal("0.00")

                if account and (debit > 0 or credit > 0):
                    total_debit += debit
                    total_credit += credit
                    valid_lines += 1

                if debit > 0 and credit > 0:
                    messages.error(request, "One line cannot have both debit and credit.")
                    return redirect("journal_edit", entry_id=entry.id)

            if valid_lines < 2:
                messages.error(request, "Journal entry must have at least 2 lines.")
            elif total_debit != total_credit:
                messages.error(
                    request,
                    f"Debit and Credit must be equal. Debit: {total_debit}, Credit: {total_credit}.",
                )
            elif total_debit <= 0:
                messages.error(request, "Total debit and credit must be more than zero.")
            else:
                with transaction.atomic():
                    form.save()
                    formset.save()

                messages.success(request, f"Journal Entry {entry.entry_no} updated successfully.")
                return redirect("journal_detail", entry_id=entry.id)

    else:
        form = JournalEntryForm(instance=entry)
        formset = JournalEntryLineFormSet(
            instance=entry,
            form_kwargs={"company": company},
        )

    return render(request, "accounting/journal_form.html", {
        "company": company,
        "form": form,
        "formset": formset,
        "entry": entry,
        "page_title": "Edit Journal Entry",
        "button_text": "Save Changes",
    })


@login_required
def journal_delete(request, entry_id):
    company, response = require_company_access(request)
    if response:
        return response

    entry = get_object_or_404(
        JournalEntry,
        id=entry_id,
        company=company,
    )

    if request.method == "POST":
        entry_no = entry.entry_no
        entry.delete()
        messages.success(request, f"Journal Entry {entry_no} deleted successfully.")
        return redirect("journal_list")

    return render(request, "accounting/journal_delete.html", {
        "company": company,
        "entry": entry,
    })


# =========================================================
# ACCOUNTING DATA SHORTCUT PAGES
# =========================================================

@login_required
def accounting_data_list(request):
    return journal_list(request)


@login_required
def accounting_data_create(request):
    return journal_create(request)


# =========================================================
# EXCEL IMPORT - JOURNAL
# =========================================================

@login_required
def accounting_import_excel(request):
    company, response = require_company_access(request)
    if response:
        return response

    has_accounts = ChartOfAccount.objects.filter(
        company=company,
        is_active=True,
        is_group=False,
    ).exists()

    if not has_accounts:
        messages.warning(request, "Please create Chart of Accounts first.")
        return redirect("chart_of_accounts")

    if request.method == "POST":
        excel_file = request.FILES.get("file")

        if not excel_file:
            messages.error(request, "Please choose an Excel file.")
            return redirect("accounting_import_excel")

        if not excel_file.name.lower().endswith(".xlsx"):
            messages.error(request, "Please upload .xlsx file only.")
            return redirect("accounting_import_excel")

        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active

            created_entries = 0
            created_lines = 0
            errors = []
            grouped_rows = {}

            for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                entry_date = parse_excel_date(row[0] if len(row) > 0 else None)
                reference_no = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                account_code = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                line_description = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                debit = to_decimal(row[4] if len(row) > 4 else 0)
                credit = to_decimal(row[5] if len(row) > 5 else 0)
                entry_description = str(row[6]).strip() if len(row) > 6 and row[6] else ""

                if not account_code and debit == 0 and credit == 0:
                    continue

                if not account_code:
                    errors.append(f"Row {row_number}: account code is missing.")
                    continue

                try:
                    account = ChartOfAccount.objects.get(
                        company=company,
                        code=account_code,
                        is_active=True,
                        is_group=False,
                    )
                except ChartOfAccount.DoesNotExist:
                    errors.append(f"Row {row_number}: account code {account_code} not found.")
                    continue

                if debit > 0 and credit > 0:
                    errors.append(f"Row {row_number}: cannot have both debit and credit.")
                    continue

                if debit == 0 and credit == 0:
                    errors.append(f"Row {row_number}: debit and credit are both zero.")
                    continue

                group_key = f"{entry_date}|{reference_no}|{entry_description}"

                if group_key not in grouped_rows:
                    grouped_rows[group_key] = {
                        "entry_date": entry_date,
                        "reference_no": reference_no,
                        "description": entry_description,
                        "lines": [],
                    }

                grouped_rows[group_key]["lines"].append({
                    "account": account,
                    "description": line_description,
                    "debit": debit,
                    "credit": credit,
                    "row_number": row_number,
                })

            with transaction.atomic():
                for group_key, data in grouped_rows.items():
                    lines = data["lines"]
                    total_debit = sum((line["debit"] for line in lines), Decimal("0.00"))
                    total_credit = sum((line["credit"] for line in lines), Decimal("0.00"))

                    if len(lines) < 2:
                        row_numbers = ", ".join(str(line["row_number"]) for line in lines)
                        errors.append(f"Rows {row_numbers}: journal must have at least 2 lines.")
                        continue

                    if total_debit != total_credit:
                        row_numbers = ", ".join(str(line["row_number"]) for line in lines)
                        errors.append(
                            f"Rows {row_numbers}: debit and credit not equal. "
                            f"Debit {total_debit}, Credit {total_credit}."
                        )
                        continue

                    entry = JournalEntry.objects.create(
                        company=company,
                        entry_date=data["entry_date"],
                        reference_no=data["reference_no"],
                        description=data["description"],
                        status=get_posted_status(),
                        created_by=request.user,
                    )

                    for line in lines:
                        JournalEntryLine.objects.create(
                            journal_entry=entry,
                            account=line["account"],
                            description=line["description"],
                            debit=line["debit"],
                            credit=line["credit"],
                        )
                        created_lines += 1

                    created_entries += 1

            for err in errors[:15]:
                messages.warning(request, err)

            if created_entries:
                messages.success(
                    request,
                    f"Imported {created_entries} journal entries and {created_lines} lines successfully.",
                )
                return redirect("journal_list")

            messages.warning(request, "No journal entries imported. Please check your Excel file.")
            return redirect("accounting_import_excel")

        except Exception as e:
            messages.error(request, f"Import failed: {e}")
            return redirect("accounting_import_excel")

    return render(request, "accounting/accounting_import_excel.html", {
        "company": company,
    })


@login_required
def download_journal_import_sample(request):
    company, response = require_company_access(request)
    if response:
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Import"

    headers = [
        "Date",
        "Reference No",
        "Account Code",
        "Line Description",
        "Debit",
        "Credit",
        "Entry Description",
    ]

    rows = [
        ["2026-05-16", "INV-001", "111500", "Cash received", 100, 0, "Sale income"],
        ["2026-05-16", "INV-001", "410000", "Sale revenue", 0, 100, "Sale income"],
        ["2026-05-16", "EXP-001", "650100", "Rental expense", 50, 0, "Office rent"],
        ["2026-05-16", "EXP-001", "111500", "Cash payment", 0, 50, "Office rent"],
    ]

    ws.append(headers)

    for row in rows:
        ws.append(row)

    blue_fill = PatternFill("solid", fgColor="0070C0")
    white_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        length = 14
        for cell in column_cells:
            if cell.value:
                length = max(length, len(str(cell.value)) + 2)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length, 35)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Journal_Import_Template.xlsx"'

    wb.save(response)
    return response


# =========================================================
# BULK UPDATE - VENDOR / ITEM / CUSTOMER
# =========================================================

ITEM_BULK_COLUMNS = [
    "ITEM", "ITEM_NAME", "ITEM_CODE", "LOCAL_NAME", "ITEM_GROUP", "ITEM_BRAND", "DESCRIPTION",
    "NEGATIVE_SALE", "FOR_PURCHASE", "FOR_SALE", "ALARM_QTY", "COST",
    "ACCOUNT_ASSET_CODE", "ACCOUNT_COGS_CODE", "ACCOUNT_REVENUE_CODE",
    "MEMO", "DETAIL_MEMO", "ACTIVE", "UNITSET_NAME",
    "PRICE_1", "BARCODE_1", "PRICE_2", "BARCODE_2", "PRICE_3", "BARCODE_3",
]

CUSTOMER_BULK_COLUMNS = [
    "CUSTOMER", "CUSTOMER_NAME", "CUSTOMER_CODE", "LOCAL_CUSTOMER_NAME", "CUSTOMER_TYPE",
    "SALE_PERSON", "REGION", "CURRENCY", "PRICE_LEVEL", "INVOICE_TYPE", "CREDIT_LIMIT",
    "CREDIT_TERM", "EMAIL", "PHONE", "VATTIN", "IS_ALLOW_OVER_CREDIT", "ACTIVE",
    "HOUSE_NO", "STREET", "COMMUNE", "DISTRICT", "CITY",
    "LOCAL_HOUSE_NO", "LOCAL_STREET", "LOCAL_COMMUNE", "LOCAL_DISTRICT", "LOCAL_CITY",
    "ADDRESS", "MEMO", "CONTACT_NAME", "CONTACT_PHONE", "CONTACT_EMAIL", "BRANCHES", "GRADE",
]

# Vendor is template only until vendor model is connected.
VENDOR_BULK_COLUMNS = [
    "VENDOR", "VENDOR_NAME", "VENDOR_CODE", "LOCAL_VENDOR_NAME",
    "REGION", "CURRENCY", "EMAIL", "PHONE", "VATTIN",
    "ACTIVE", "ADDRESS", "MEMO",
]


def bulk_clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def bulk_clean_decimal(value):
    if value is None or str(value).strip() == "":
        return Decimal("0.00")

    try:
        return Decimal(str(value).replace(",", "").strip()).quantize(Decimal("0.01"))
    except Exception:
        raise ValueError("must be number")


def bulk_clean_int(value):
    if value is None or str(value).strip() == "":
        return 0

    try:
        return int(float(str(value).replace(",", "").strip()))
    except Exception:
        raise ValueError("must be integer")


def bulk_clean_bool(value):
    if value is None or str(value).strip() == "":
        return False

    text = str(value).strip().lower()

    if text in ["true", "yes", "1", "y", "active"]:
        return True

    if text in ["false", "no", "0", "n", "inactive"]:
        return False

    raise ValueError("must be TRUE or FALSE")


def bulk_get_headers(sheet):
    return [bulk_clean_text(cell.value) for cell in sheet[1]]


def bulk_validate_columns(sheet, expected_columns, sheet_name):
    headers = bulk_get_headers(sheet)
    errors = []

    for col in expected_columns:
        if col not in headers:
            errors.append({
                "row": 1,
                "type": f"{sheet_name} Column",
                "code": "-",
                "error": f"Missing column: {col}",
            })

    for col in headers:
        if col and col not in expected_columns:
            errors.append({
                "row": 1,
                "type": f"{sheet_name} Column",
                "code": "-",
                "error": f"Wrong column: {col}",
            })

    return errors


def bulk_row_to_dict(sheet, row_number, expected_columns):
    headers = bulk_get_headers(sheet)
    row_cells = list(sheet[row_number])

    data = {}

    for col in expected_columns:
        if col in headers:
            index = headers.index(col)
            data[col] = row_cells[index].value if index < len(row_cells) else None
        else:
            data[col] = None

    return data


def get_bulk_page_data(bulk_type):
    if bulk_type == "vendors":
        return {
            "bulk_type": "vendors",
            "page_title": "Update Vendors",
            "page_icon": "🏪",
            "page_note": "Vendor bulk update page. Vendor Excel import will be connected after vendor model is connected.",
            "download_label": "Download Vendor Sample",
            "upload_label": "Upload Vendor Excel",
            "upload_enabled": False,
            "sample_type": "vendors",
        }

    if bulk_type == "customers":
        return {
            "bulk_type": "customers",
            "page_title": "Update Customers",
            "page_icon": "👥",
            "page_note": "Download customer sample, upload Excel, check columns, then update customer data.",
            "download_label": "Download Customer Sample",
            "upload_label": "Upload Customer Excel",
            "upload_enabled": True,
            "sample_type": "customers",
        }

    return {
        "bulk_type": "items",
        "page_title": "Update Items",
        "page_icon": "📦",
        "page_note": "Download item sample, upload Excel, check columns, then update item data.",
        "download_label": "Download Item Sample",
        "upload_label": "Upload Item Excel",
        "upload_enabled": True,
        "sample_type": "items",
    }


@login_required
def bulk_update_vendors(request):
    return bulk_update_menu(request, bulk_type="vendors")


@login_required
def bulk_update_items(request):
    return bulk_update_menu(request, bulk_type="items")


@login_required
def bulk_update_customers(request):
    return bulk_update_menu(request, bulk_type="customers")


@login_required
def bulk_update_menu(request, bulk_type="items"):
    company, response = require_company_access(request)
    if response:
        return response

    page_data = get_bulk_page_data(bulk_type)

    logs = BulkImportLog.objects.filter(company=company).order_by("-created_at")[:30]
    total_items = AccountItem.objects.filter(company=company).count()
    total_customers = AccountCustomer.objects.filter(company=company).count()

    return render(request, "accounting/bulk_update_menu.html", {
        "company": company,
        "logs": logs,
        "total_items": total_items,
        "total_customers": total_customers,
        "bulk_type": page_data["bulk_type"],
        "page_title": page_data["page_title"],
        "page_icon": page_data["page_icon"],
        "page_note": page_data["page_note"],
        "download_label": page_data["download_label"],
        "upload_label": page_data["upload_label"],
        "upload_enabled": page_data["upload_enabled"],
        "sample_type": page_data["sample_type"],
    })


def style_bulk_workbook(wb):
    blue_fill = PatternFill("solid", fgColor="1D4ED8")
    white_font = Font(color="FFFFFF", bold=True)

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"

        for cell in sheet[1]:
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")

        for col in sheet.columns:
            max_length = 12
            col_letter = col[0].column_letter

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)) + 2)

            sheet.column_dimensions[col_letter].width = min(max_length, 30)


@login_required
def download_bulk_update_sample(request):
    company, response = require_company_access(request)
    if response:
        return response

    bulk_type = (request.GET.get("type") or "all").strip().lower()

    wb = Workbook()

    # Items sheet
    ws = wb.active
    ws.title = "Items"
    ws.append(ITEM_BULK_COLUMNS)

    item_rows = [
        [
            "fanta", "fanta", "F-001", "", "ទឹកក្រូច", "", "",
            True, True, True, 100, 1,
            "131000", "510000", "410000",
            "", "", True, "case/2L",
            1, "", 12, "", 24, "",
        ],
        [
            "cocacola", "cocacola", "CO-001", "", "ទឹកក្រូច", "", "",
            True, True, True, 100, 1,
            "131000", "510000", "410000",
            "", "", True, "case/2L",
            1.5, "", 12, "", 36, "",
        ],
    ]

    for row in item_rows:
        ws.append(row)

    # Customers sheet
    ws2 = wb.create_sheet("Customers")
    ws2.append(CUSTOMER_BULK_COLUMNS)

    customer_rows = [
        [
            "Customer 1", "Customer 1", "C1", "អតិថិន១", "",
            "", "phnom penh", "Dollar USD", "", "TAX_INVOICE", 0,
            0, "", "", "K001-123456789", True, True,
            "E1122", "#2004", "Commune", "District", "City",
            "House Number", "#2004", "Khmer Commune", "Khmer District", "Phnom Penh",
            "", "", "", "", "", "PP", "",
        ],
        [
            "Customer 2", "Customer 2", "C2", "អតិថិន២", "",
            "", "phnom penh", "Riel", "", "TAX_INVOICE", 0,
            0, "", "", "K001-123456790", True, True,
            "E1123", "#2005", "Commune", "District", "City",
            "House Number", "#2005", "Khmer Commune", "Khmer District", "Phnom Penh",
            "", "", "", "", "", "KD", "",
        ],
    ]

    for row in customer_rows:
        ws2.append(row)

    # Vendors placeholder sheet
    ws3 = wb.create_sheet("Vendors")
    ws3.append(VENDOR_BULK_COLUMNS)
    ws3.append([
        "Vendor 1", "Vendor 1", "V1", "អ្នកផ្គត់ផ្គង់១",
        "phnom penh", "Dollar USD", "", "012345678", "K001-123456789",
        True, "Phnom Penh", "",
    ])

    if bulk_type == "items":
        del wb["Customers"]
        del wb["Vendors"]
        filename = "Bulk_Update_Items_Template.xlsx"
    elif bulk_type == "customers":
        del wb["Items"]
        del wb["Vendors"]
        filename = "Bulk_Update_Customers_Template.xlsx"
    elif bulk_type == "vendors":
        del wb["Items"]
        del wb["Customers"]
        filename = "Bulk_Update_Vendors_Template.xlsx"
    else:
        filename = "Bulk_Update_Item_Customer_Template.xlsx"

    style_bulk_workbook(wb)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def upload_bulk_update(request):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method != "POST":
        return redirect("bulk_update_items")

    bulk_type = (request.POST.get("bulk_type") or "items").strip().lower()

    if bulk_type == "vendors":
        messages.warning(request, "Vendor bulk upload is not connected yet. Add vendor model first.")
        return redirect("bulk_update_vendors")

    excel_file = request.FILES.get("excel_file")

    if not excel_file:
        messages.error(request, "Please choose Excel file first.")
        if bulk_type == "customers":
            return redirect("bulk_update_customers")
        return redirect("bulk_update_items")

    if not excel_file.name.lower().endswith(".xlsx"):
        messages.error(request, "Only .xlsx Excel file is allowed.")
        if bulk_type == "customers":
            return redirect("bulk_update_customers")
        return redirect("bulk_update_items")

    errors = []
    item_created = 0
    item_updated = 0
    customer_created = 0
    customer_updated = 0

    try:
        wb = load_workbook(excel_file, data_only=True)
    except Exception:
        messages.error(request, "Cannot read Excel file. Please download sample and fill again.")
        if bulk_type == "customers":
            return redirect("bulk_update_customers")
        return redirect("bulk_update_items")

    # Check sheet by type
    if bulk_type == "items":
        if "Items" not in wb.sheetnames:
            errors.append({
                "row": "-",
                "type": "Sheet",
                "code": "-",
                "error": "Missing sheet: Items",
            })
        else:
            item_sheet = wb["Items"]
            errors.extend(bulk_validate_columns(item_sheet, ITEM_BULK_COLUMNS, "Item"))

    elif bulk_type == "customers":
        if "Customers" not in wb.sheetnames:
            errors.append({
                "row": "-",
                "type": "Sheet",
                "code": "-",
                "error": "Missing sheet: Customers",
            })
        else:
            customer_sheet = wb["Customers"]
            errors.extend(bulk_validate_columns(customer_sheet, CUSTOMER_BULK_COLUMNS, "Customer"))

    else:
        if "Items" not in wb.sheetnames:
            errors.append({
                "row": "-",
                "type": "Sheet",
                "code": "-",
                "error": "Missing sheet: Items",
            })

        if "Customers" not in wb.sheetnames:
            errors.append({
                "row": "-",
                "type": "Sheet",
                "code": "-",
                "error": "Missing sheet: Customers",
            })

        if not errors:
            item_sheet = wb["Items"]
            customer_sheet = wb["Customers"]
            errors.extend(bulk_validate_columns(item_sheet, ITEM_BULK_COLUMNS, "Item"))
            errors.extend(bulk_validate_columns(customer_sheet, CUSTOMER_BULK_COLUMNS, "Customer"))

    if errors:
        log = BulkImportLog.objects.create(
            company=company,
            uploaded_by=request.user,
            file_name=excel_file.name,
            status=BulkImportLog.STATUS_FAILED,
            error_count=len(errors),
            error_report=errors,
        )

        messages.error(request, "Upload failed. Excel sheet or column is wrong.")
        return render(request, "accounting/bulk_update_result.html", {
            "company": company,
            "log": log,
            "errors": errors,
            "item_created": item_created,
            "item_updated": item_updated,
            "customer_created": customer_created,
            "customer_updated": customer_updated,
            "status": BulkImportLog.STATUS_FAILED,
            "bulk_type": bulk_type,
        })

    with transaction.atomic():
        # Import items only for item/all
        if bulk_type in ["items", "all"]:
            item_sheet = wb["Items"]

            for row_number in range(2, item_sheet.max_row + 1):
                data = bulk_row_to_dict(item_sheet, row_number, ITEM_BULK_COLUMNS)

                item_code = bulk_clean_text(data.get("ITEM_CODE"))
                item_name = bulk_clean_text(data.get("ITEM_NAME"))

                if not item_code and not item_name:
                    continue

                try:
                    if not item_code:
                        raise ValueError("ITEM_CODE is required")

                    if not item_name:
                        raise ValueError("ITEM_NAME is required")

                    obj, created = AccountItem.objects.update_or_create(
                        company=company,
                        item_code=item_code,
                        defaults={
                            "item": bulk_clean_text(data.get("ITEM")) or item_name,
                            "item_name": item_name,
                            "local_name": bulk_clean_text(data.get("LOCAL_NAME")),
                            "item_group": bulk_clean_text(data.get("ITEM_GROUP")),
                            "item_brand": bulk_clean_text(data.get("ITEM_BRAND")),
                            "description": bulk_clean_text(data.get("DESCRIPTION")),
                            "negative_sale": bulk_clean_bool(data.get("NEGATIVE_SALE")),
                            "for_purchase": bulk_clean_bool(data.get("FOR_PURCHASE")),
                            "for_sale": bulk_clean_bool(data.get("FOR_SALE")),
                            "alarm_qty": bulk_clean_decimal(data.get("ALARM_QTY")),
                            "cost": bulk_clean_decimal(data.get("COST")),
                            "account_asset_code": bulk_clean_text(data.get("ACCOUNT_ASSET_CODE")),
                            "account_cogs_code": bulk_clean_text(data.get("ACCOUNT_COGS_CODE")),
                            "account_revenue_code": bulk_clean_text(data.get("ACCOUNT_REVENUE_CODE")),
                            "memo": bulk_clean_text(data.get("MEMO")),
                            "detail_memo": bulk_clean_text(data.get("DETAIL_MEMO")),
                            "active": bulk_clean_bool(data.get("ACTIVE")),
                            "unitset_name": bulk_clean_text(data.get("UNITSET_NAME")),
                            "price_1": bulk_clean_decimal(data.get("PRICE_1")),
                            "barcode_1": bulk_clean_text(data.get("BARCODE_1")),
                            "price_2": bulk_clean_decimal(data.get("PRICE_2")),
                            "barcode_2": bulk_clean_text(data.get("BARCODE_2")),
                            "price_3": bulk_clean_decimal(data.get("PRICE_3")),
                            "barcode_3": bulk_clean_text(data.get("BARCODE_3")),
                        },
                    )

                    if created:
                        item_created += 1
                    else:
                        item_updated += 1

                except Exception as e:
                    errors.append({
                        "row": row_number,
                        "type": "Item",
                        "code": item_code or "-",
                        "error": str(e),
                    })

        # Import customers only for customer/all
        if bulk_type in ["customers", "all"]:
            customer_sheet = wb["Customers"]

            for row_number in range(2, customer_sheet.max_row + 1):
                data = bulk_row_to_dict(customer_sheet, row_number, CUSTOMER_BULK_COLUMNS)

                customer_code = bulk_clean_text(data.get("CUSTOMER_CODE"))
                customer_name = bulk_clean_text(data.get("CUSTOMER_NAME"))

                if not customer_code and not customer_name:
                    continue

                try:
                    if not customer_code:
                        raise ValueError("CUSTOMER_CODE is required")

                    if not customer_name:
                        raise ValueError("CUSTOMER_NAME is required")

                    obj, created = AccountCustomer.objects.update_or_create(
                        company=company,
                        customer_code=customer_code,
                        defaults={
                            "customer": bulk_clean_text(data.get("CUSTOMER")) or customer_name,
                            "customer_name": customer_name,
                            "local_customer_name": bulk_clean_text(data.get("LOCAL_CUSTOMER_NAME")),
                            "customer_type": bulk_clean_text(data.get("CUSTOMER_TYPE")),
                            "sale_person": bulk_clean_text(data.get("SALE_PERSON")),
                            "region": bulk_clean_text(data.get("REGION")),
                            "currency": bulk_clean_text(data.get("CURRENCY")),
                            "price_level": bulk_clean_text(data.get("PRICE_LEVEL")),
                            "invoice_type": bulk_clean_text(data.get("INVOICE_TYPE")),
                            "credit_limit": bulk_clean_decimal(data.get("CREDIT_LIMIT")),
                            "credit_term": bulk_clean_int(data.get("CREDIT_TERM")),
                            "email": bulk_clean_text(data.get("EMAIL")),
                            "phone": bulk_clean_text(data.get("PHONE")),
                            "vattin": bulk_clean_text(data.get("VATTIN")),
                            "is_allow_over_credit": bulk_clean_bool(data.get("IS_ALLOW_OVER_CREDIT")),
                            "active": bulk_clean_bool(data.get("ACTIVE")),
                            "house_no": bulk_clean_text(data.get("HOUSE_NO")),
                            "street": bulk_clean_text(data.get("STREET")),
                            "commune": bulk_clean_text(data.get("COMMUNE")),
                            "district": bulk_clean_text(data.get("DISTRICT")),
                            "city": bulk_clean_text(data.get("CITY")),
                            "local_house_no": bulk_clean_text(data.get("LOCAL_HOUSE_NO")),
                            "local_street": bulk_clean_text(data.get("LOCAL_STREET")),
                            "local_commune": bulk_clean_text(data.get("LOCAL_COMMUNE")),
                            "local_district": bulk_clean_text(data.get("LOCAL_DISTRICT")),
                            "local_city": bulk_clean_text(data.get("LOCAL_CITY")),
                            "address": bulk_clean_text(data.get("ADDRESS")),
                            "memo": bulk_clean_text(data.get("MEMO")),
                            "contact_name": bulk_clean_text(data.get("CONTACT_NAME")),
                            "contact_phone": bulk_clean_text(data.get("CONTACT_PHONE")),
                            "contact_email": bulk_clean_text(data.get("CONTACT_EMAIL")),
                            "branches": bulk_clean_text(data.get("BRANCHES")),
                            "grade": bulk_clean_text(data.get("GRADE")),
                        },
                    )

                    if created:
                        customer_created += 1
                    else:
                        customer_updated += 1

                except Exception as e:
                    errors.append({
                        "row": row_number,
                        "type": "Customer",
                        "code": customer_code or "-",
                        "error": str(e),
                    })

        status = BulkImportLog.STATUS_FAILED if errors else BulkImportLog.STATUS_SUCCESS

        log = BulkImportLog.objects.create(
            company=company,
            uploaded_by=request.user,
            file_name=excel_file.name,
            status=status,
            item_created=item_created,
            item_updated=item_updated,
            customer_created=customer_created,
            customer_updated=customer_updated,
            error_count=len(errors),
            error_report=errors,
        )

    if errors:
        messages.error(request, f"Import finished with {len(errors)} error(s). Please check report.")
    else:
        messages.success(request, "Bulk update imported successfully.")

    return render(request, "accounting/bulk_update_result.html", {
        "company": company,
        "log": log,
        "errors": errors,
        "item_created": item_created,
        "item_updated": item_updated,
        "customer_created": customer_created,
        "customer_updated": customer_updated,
        "status": status,
        "bulk_type": bulk_type,
    })


@login_required
def bulk_update_report(request, log_id):
    company, response = require_company_access(request)
    if response:
        return response

    log = get_object_or_404(
        BulkImportLog,
        id=log_id,
        company=company,
    )

    return render(request, "accounting/bulk_update_report.html", {
        "company": company,
        "log": log,
        "errors": log.error_report or [],
    })


# =========================================================
# AUTO REPORTS
# =========================================================

@login_required
def accounting_reports(request):
    company, response = require_company_access(request)
    if response:
        return response

    today = timezone.localdate()
    date_from = (request.GET.get("date_from") or today.replace(day=1).strftime("%Y-%m-%d")).strip()
    date_to = (request.GET.get("date_to") or today.strftime("%Y-%m-%d")).strip()
    posted_status = get_posted_status()

    # =====================================================
    # Base posted journal lines
    # =====================================================
    lines = JournalEntryLine.objects.filter(
        journal_entry__company=company,
        journal_entry__status=posted_status,
    ).select_related("journal_entry", "account")

    if date_from:
        lines = lines.filter(journal_entry__entry_date__gte=date_from)

    if date_to:
        lines = lines.filter(journal_entry__entry_date__lte=date_to)

    # =====================================================
    # Trial Balance / P&L / Balance Sheet account summary
    # =====================================================
    account_rows = (
        lines
        .values(
            "account_id",
            "account__code",
            "account__name",
            "account__account_type",
            "account__report_type",
            "account__report_section",
            "account__normal_balance",
        )
        .annotate(
            debit=Sum("debit"),
            credit=Sum("credit"),
        )
        .order_by("account__code")
    )

    trial_balance_rows = []
    profit_loss_rows = []
    balance_sheet_rows = []

    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")

    total_revenue = Decimal("0.00")
    total_cogs = Decimal("0.00")
    total_expense = Decimal("0.00")
    total_other_income = Decimal("0.00")
    total_other_expense = Decimal("0.00")

    total_assets = Decimal("0.00")
    total_liabilities = Decimal("0.00")
    total_equity = Decimal("0.00")

    for row in account_rows:
        debit = row["debit"] or Decimal("0.00")
        credit = row["credit"] or Decimal("0.00")
        account_type = row["account__account_type"]
        report_type = row["account__report_type"]

        total_debit += debit
        total_credit += credit

        if account_type in ["asset", "expense", "cogs", "other_expense"]:
            balance = debit - credit
        else:
            balance = credit - debit

        data = {
            "account_id": row["account_id"],
            "code": row["account__code"],
            "name": row["account__name"],
            "account_type": account_type,
            "report_type": report_type,
            "report_section": row["account__report_section"],
            "normal_balance": row["account__normal_balance"],
            "debit": debit,
            "credit": credit,
            "balance": balance,
        }

        trial_balance_rows.append(data)

        if report_type == ChartOfAccount.REPORT_PROFIT_LOSS:
            profit_loss_rows.append(data)

            if account_type == ChartOfAccount.ACCOUNT_TYPE_REVENUE:
                total_revenue += balance
            elif account_type == ChartOfAccount.ACCOUNT_TYPE_COGS:
                total_cogs += balance
            elif account_type == ChartOfAccount.ACCOUNT_TYPE_EXPENSE:
                total_expense += balance
            elif account_type == ChartOfAccount.ACCOUNT_TYPE_OTHER_INCOME:
                total_other_income += balance
            elif account_type == ChartOfAccount.ACCOUNT_TYPE_OTHER_EXPENSE:
                total_other_expense += balance

        if report_type == ChartOfAccount.REPORT_BALANCE_SHEET:
            balance_sheet_rows.append(data)

            if account_type == ChartOfAccount.ACCOUNT_TYPE_ASSET:
                total_assets += balance
            elif account_type == ChartOfAccount.ACCOUNT_TYPE_LIABILITY:
                total_liabilities += balance
            elif account_type == ChartOfAccount.ACCOUNT_TYPE_EQUITY:
                total_equity += balance

    gross_profit = total_revenue - total_cogs
    net_profit = gross_profit - total_expense + total_other_income - total_other_expense
    balance_check = total_assets - (total_liabilities + total_equity + net_profit)

    # =====================================================
    # General Ledger / Journal Report preview rows
    # Click row -> journal detail
    # =====================================================
    journal_report_rows = (
        lines
        .order_by("journal_entry__entry_date", "journal_entry_id", "id")[:300]
    )

    general_ledger_rows = journal_report_rows

    # =====================================================
    # Profit/Loss by Month preview
    # Click account -> ledger detail
    # =====================================================
    pl_month_rows = []
    year = today.year

    try:
        if date_from:
            year = datetime.strptime(date_from, "%Y-%m-%d").year
    except Exception:
        year = today.year

    pl_accounts = ChartOfAccount.objects.filter(
        company=company,
        is_active=True,
        is_group=False,
        report_type=ChartOfAccount.REPORT_PROFIT_LOSS,
    ).order_by("code")

    for account in pl_accounts:
        monthly_amounts = []
        row_total = Decimal("0.00")

        for month in range(1, 13):
            month_lines = JournalEntryLine.objects.filter(
                journal_entry__company=company,
                journal_entry__status=posted_status,
                journal_entry__entry_date__year=year,
                journal_entry__entry_date__month=month,
                account=account,
            )

            sums = month_lines.aggregate(
                debit=Sum("debit"),
                credit=Sum("credit"),
            )

            debit = sums["debit"] or Decimal("0.00")
            credit = sums["credit"] or Decimal("0.00")

            if account.account_type in [
                ChartOfAccount.ACCOUNT_TYPE_REVENUE,
                ChartOfAccount.ACCOUNT_TYPE_OTHER_INCOME,
            ]:
                amount = credit - debit
            else:
                amount = debit - credit

            monthly_amounts.append(amount)
            row_total += amount

        if row_total != 0:
            pl_month_rows.append({
                "account_id": account.id,
                "code": account.code,
                "name": account.name,
                "months": monthly_amounts,
                "total": row_total,
            })

    # =====================================================
    # AP Aging preview
    # Click row -> vendor transaction detail
    # =====================================================
    ap_aging_rows = []
    try:
        from vendors.models import VendorTransaction

        ap_qs = VendorTransaction.objects.filter(
            company=company,
            status=VendorTransaction.STATUS_POSTED,
            transaction_type__in=[
                VendorTransaction.TYPE_PURCHASE_ORDER,
                VendorTransaction.TYPE_ADJUSTMENT,
            ],
        ).select_related("vendor").order_by("vendor__name", "transaction_date")

        if date_to:
            ap_qs = ap_qs.filter(transaction_date__lte=date_to)

        for tx in ap_qs[:300]:
            age = (today - tx.transaction_date).days
            amount = tx.amount or Decimal("0.00")

            ap_aging_rows.append({
                "id": tx.id,
                "vendor": tx.vendor.name if tx.vendor else "-",
                "date": tx.transaction_date,
                "number": tx.number,
                "type": tx.get_transaction_type_display(),
                "current": amount if age <= 0 else Decimal("0.00"),
                "days_1_30": amount if 1 <= age <= 30 else Decimal("0.00"),
                "days_31_60": amount if 31 <= age <= 60 else Decimal("0.00"),
                "days_61_90": amount if 61 <= age <= 90 else Decimal("0.00"),
                "over_90": amount if age > 90 else Decimal("0.00"),
            })
    except Exception:
        ap_aging_rows = []

    # =====================================================
    # AR Aging preview
    # Click row -> customer transaction detail
    # =====================================================
    ar_aging_rows = []
    try:
        from customers.models import CustomerTransaction

        ar_qs = CustomerTransaction.objects.filter(
            company=company,
            status=CustomerTransaction.STATUS_POSTED,
            transaction_type__in=[
                CustomerTransaction.TYPE_INVOICE,
                CustomerTransaction.TYPE_ADJUSTMENT,
            ],
        ).select_related("customer").order_by("customer__name", "transaction_date")

        if date_to:
            ar_qs = ar_qs.filter(transaction_date__lte=date_to)

        for tx in ar_qs[:300]:
            age = (today - tx.transaction_date).days
            amount = tx.amount or Decimal("0.00")

            ar_aging_rows.append({
                "id": tx.id,
                "customer": tx.customer.name,
                "date": tx.transaction_date,
                "number": tx.number,
                "type": tx.get_transaction_type_display(),
                "current": amount if age <= 0 else Decimal("0.00"),
                "days_1_30": amount if 1 <= age <= 30 else Decimal("0.00"),
                "days_31_60": amount if 31 <= age <= 60 else Decimal("0.00"),
                "days_61_90": amount if 61 <= age <= 90 else Decimal("0.00"),
                "over_90": amount if age > 90 else Decimal("0.00"),
            })
    except Exception:
        ar_aging_rows = []

    # =====================================================
    # Cash Flow preview
    # Click account -> ledger detail
    # =====================================================
    cash_flow_rows = []
    cash_accounts = ChartOfAccount.objects.filter(
        company=company,
        is_active=True,
        is_group=False,
        account_type=ChartOfAccount.ACCOUNT_TYPE_ASSET,
        name__icontains="cash",
    ).order_by("code")

    total_cash_change = Decimal("0.00")

    for account in cash_accounts:
        cash_lines = lines.filter(account=account)
        sums = cash_lines.aggregate(
            debit=Sum("debit"),
            credit=Sum("credit"),
        )

        debit = sums["debit"] or Decimal("0.00")
        credit = sums["credit"] or Decimal("0.00")
        amount = debit - credit

        if amount != 0:
            total_cash_change += amount
            cash_flow_rows.append({
                "account_id": account.id,
                "code": account.code,
                "name": account.name,
                "section": account.report_section,
                "amount": amount,
            })

    return render(request, "accounting/accounting_reports.html", {
        "company": company,
        "date_from": date_from,
        "date_to": date_to,
        "report_year": year,

        "trial_balance_rows": trial_balance_rows,
        "profit_loss_rows": profit_loss_rows,
        "balance_sheet_rows": balance_sheet_rows,
        "pl_month_rows": pl_month_rows,
        "general_ledger_rows": general_ledger_rows,
        "journal_report_rows": journal_report_rows,
        "ap_aging_rows": ap_aging_rows,
        "ar_aging_rows": ar_aging_rows,
        "cash_flow_rows": cash_flow_rows,

        "total_debit": total_debit,
        "total_credit": total_credit,

        "total_revenue": total_revenue,
        "total_cogs": total_cogs,
        "gross_profit": gross_profit,
        "total_expense": total_expense,
        "total_other_income": total_other_income,
        "total_other_expense": total_other_expense,
        "net_profit": net_profit,

        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "total_equity": total_equity,
        "balance_check": balance_check,

        "total_cash_change": total_cash_change,
    })
# =========================================================
# ACCOUNTING MENU PAGES
# =========================================================

@login_required
def report_mapping(request):
    company, response = require_company_access(request)
    if response:
        return response

    return render(request, "accounting/report_mapping.html", {
        "company": company,
    })


@login_required
def banking_menu(request):
    company, response = require_company_access(request)
    if response:
        return response

    return render(request, "accounting/banking_menu.html", {
        "company": company,
    })


@login_required
def bank_deposit(request):
    company, response = require_company_access(request)
    if response:
        return response

    return render(request, "accounting/bank_deposit.html", {
        "company": company,
    })


@login_required
def landed_cost_allocation(request):
    company, response = require_company_access(request)
    if response:
        return response

    return render(request, "accounting/landed_cost_allocation.html", {
        "company": company,
    })


@login_required
def find_transaction(request):
    company, response = require_company_access(request)
    if response:
        return response

    query = (request.GET.get("q") or "").strip()

    entries = JournalEntry.objects.filter(company=company)

    if query:
        entries = entries.filter(
            Q(reference_no__icontains=query)
            | Q(description__icontains=query)
            | Q(lines__account__code__icontains=query)
            | Q(lines__account__name__icontains=query)
        ).distinct()

    entries = entries.prefetch_related("lines", "lines__account").order_by("-entry_date", "-id")[:100]

    return render(request, "accounting/find_transaction.html", {
        "company": company,
        "query": query,
        "entries": entries,
    })


@login_required
def batch_transaction(request):
    company, response = require_company_access(request)
    if response:
        return response

    entries = (
        JournalEntry.objects
        .filter(company=company)
        .prefetch_related("lines", "lines__account")
        .order_by("-entry_date", "-id")[:100]
    )

    return render(request, "accounting/batch_transaction.html", {
        "company": company,
        "entries": entries,
    })


@login_required
def import_menu(request):
    company, response = require_company_access(request)
    if response:
        return response

    return render(request, "accounting/import_menu.html", {
        "company": company,
    })


@login_required
def report_ledger_detail(request, account_id):
    company, response = require_company_access(request)
    if response:
        return response

    account = get_object_or_404(
        ChartOfAccount,
        id=account_id,
        company=company,
    )

    today = timezone.localdate()
    date_from = (request.GET.get("date_from") or today.replace(day=1).strftime("%Y-%m-%d")).strip()
    date_to = (request.GET.get("date_to") or today.strftime("%Y-%m-%d")).strip()

    lines = JournalEntryLine.objects.filter(
        journal_entry__company=company,
        journal_entry__status=get_posted_status(),
        account=account,
    ).select_related(
        "journal_entry",
        "account",
    )

    if date_from:
        lines = lines.filter(journal_entry__entry_date__gte=date_from)

    if date_to:
        lines = lines.filter(journal_entry__entry_date__lte=date_to)

    lines = lines.order_by("journal_entry__entry_date", "journal_entry_id", "id")

    ledger_rows = []
    running_balance = Decimal("0.00")
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")

    for line in lines:
        debit = line.debit or Decimal("0.00")
        credit = line.credit or Decimal("0.00")

        total_debit += debit
        total_credit += credit

        if account.normal_balance == ChartOfAccount.NORMAL_CREDIT:
            running_balance += credit - debit
        else:
            running_balance += debit - credit

        ledger_rows.append({
            "line": line,
            "entry": line.journal_entry,
            "date": line.journal_entry.entry_date,
            "entry_no": line.journal_entry.entry_no,
            "reference_no": line.journal_entry.reference_no,
            "description": line.description or line.journal_entry.description,
            "debit": debit,
            "credit": credit,
            "balance": running_balance,
        })

    return render(request, "accounting/report_ledger_detail.html", {
        "company": company,
        "account": account,
        "date_from": date_from,
        "date_to": date_to,
        "ledger_rows": ledger_rows,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "running_balance": running_balance,
    })