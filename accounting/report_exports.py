from calendar import month_abbr
from datetime import date, datetime
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .models import ChartOfAccount, JournalEntry, JournalEntryLine
from .views import require_company_access


# =========================================================
# BASIC HELPERS
# =========================================================

MONEY_ZERO = Decimal("0.00")


def d(value):
    if value is None:
        return MONEY_ZERO
    return Decimal(value).quantize(Decimal("0.01"))


def parse_date(value, default=None):
    if not value:
        return default

    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return default


def money(value):
    return d(value)


def fmt_money(value):
    try:
        return f"{money(value):,.2f}"
    except Exception:
        return "0.00"


def safe_text(value, max_length=90):
    if value is None:
        return ""

    text = str(value).replace("\n", " ").replace("\r", " ").strip()

    if len(text) > max_length:
        return text[:max_length - 3] + "..."

    return text


def get_date_range(request):
    today = timezone.localdate()

    default_from = date(today.year, 1, 1)
    default_to = today

    date_from = parse_date(request.GET.get("date_from"), default_from)
    date_to = parse_date(request.GET.get("date_to"), default_to)

    return date_from, date_to


def report_subtitle(date_from=None, date_to=None, as_of=False):
    if as_of and date_to:
        return f"As of {date_to.strftime('%d-%m-%Y')}"

    if date_from and date_to:
        return f"From {date_from.strftime('%d-%m-%Y')} To {date_to.strftime('%d-%m-%Y')}"

    return ""


def posted_lines(company, date_from=None, date_to=None):
    qs = JournalEntryLine.objects.filter(
        journal_entry__company=company,
        journal_entry__status=JournalEntry.STATUS_POSTED,
    ).select_related("journal_entry", "account")

    if date_from:
        qs = qs.filter(journal_entry__entry_date__gte=date_from)

    if date_to:
        qs = qs.filter(journal_entry__entry_date__lte=date_to)

    return qs


def all_active_accounts(company):
    return ChartOfAccount.objects.filter(
        company=company,
        is_active=True,
    ).order_by("code", "name")


def detail_accounts(company):
    return all_active_accounts(company).filter(is_group=False)


def account_signed_balance(account, debit, credit):
    debit = d(debit)
    credit = d(credit)

    if account.normal_balance == ChartOfAccount.NORMAL_CREDIT:
        return credit - debit

    return debit - credit


def pl_signed_balance(account, debit, credit):
    debit = d(debit)
    credit = d(credit)

    if account.account_type in [
        ChartOfAccount.ACCOUNT_TYPE_REVENUE,
        ChartOfAccount.ACCOUNT_TYPE_OTHER_INCOME,
    ]:
        return credit - debit

    return debit - credit


def get_account_totals(company, date_from=None, date_to=None):
    lines = posted_lines(company, date_from, date_to)

    totals = (
        lines.values("account_id")
        .annotate(
            debit_total=Sum("debit"),
            credit_total=Sum("credit"),
        )
    )

    return {
        row["account_id"]: {
            "debit": d(row["debit_total"]),
            "credit": d(row["credit_total"]),
        }
        for row in totals
    }


# =========================================================
# EXCEL STYLE
# =========================================================

def make_workbook(title):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    return wb, ws


def style_title(ws, row, last_col):
    fill = PatternFill("solid", fgColor="073763")
    font = Font(color="FFFFFF", bold=True, size=14)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)

    cell = ws.cell(row=row, column=1)
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[row].height = 26


def style_header(ws, row, last_col):
    fill = PatternFill("solid", fgColor="D9EDF8")
    font = Font(color="073763", bold=True)
    border = Border(bottom=Side(style="thin", color="B7D7E8"))

    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_money_columns(ws, columns):
    for col in columns:
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = '#,##0.00;[Red]-#,##0.00;-'


def auto_width(ws):
    for column_cells in ws.columns:
        max_length = 10
        col_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)) + 2)

        ws.column_dimensions[col_letter].width = min(max_length, 45)


def write_response(wb, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def add_report_title(ws, company, title, date_from=None, date_to=None, last_col=6):
    ws["A1"] = company.name
    style_title(ws, 1, last_col)

    ws["A2"] = title
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws["A2"].font = Font(bold=True, size=13)
    ws["A2"].alignment = Alignment(horizontal="center")

    if date_from and date_to:
        ws["A3"] = f"From {date_from.strftime('%d-%m-%Y')} To {date_to.strftime('%d-%m-%Y')}"
    elif date_to:
        ws["A3"] = f"As of {date_to.strftime('%d-%m-%Y')}"
    else:
        ws["A3"] = ""

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    ws["A3"].alignment = Alignment(horizontal="center")


# =========================================================
# PDF STYLE
# =========================================================

def write_pdf_response(filename):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def simple_pdf_report(company, title, subtitle, headers, rows, filename, landscape_page=True):
    response = write_pdf_response(filename)
    page_size = landscape(A4) if landscape_page else A4

    doc = SimpleDocTemplate(
        response,
        pagesize=page_size,
        rightMargin=16,
        leftMargin=16,
        topMargin=18,
        bottomMargin=18,
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"<b>{safe_text(company.name, 120)}</b>", styles["Title"]))
    elements.append(Paragraph(f"<b>{safe_text(title, 120)}</b>", styles["Heading2"]))

    if subtitle:
        elements.append(Paragraph(safe_text(subtitle, 120), styles["Normal"]))

    elements.append(Spacer(1, 10))

    clean_headers = [safe_text(h, 40) for h in headers]
    clean_rows = []

    for row in rows:
        clean_rows.append([safe_text(value, 75) for value in row])

    table_data = [clean_headers] + clean_rows

    if not clean_rows:
        table_data.append(["No data"] + [""] * (len(clean_headers) - 1))

    table = Table(table_data, repeatRows=1)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#073763")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9EDF8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6FBFF")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))

    elements.append(table)
    doc.build(elements)

    return response


# =========================================================
# COMMON DATA BUILDERS
# =========================================================

def build_trial_balance_rows(company, date_from, date_to):
    totals = get_account_totals(company, date_from, date_to)

    rows = []
    total_debit = MONEY_ZERO
    total_credit = MONEY_ZERO

    for account in detail_accounts(company):
        row_total = totals.get(account.id, {"debit": MONEY_ZERO, "credit": MONEY_ZERO})
        debit = row_total["debit"]
        credit = row_total["credit"]
        balance = account_signed_balance(account, debit, credit)

        if debit == 0 and credit == 0:
            continue

        total_debit += debit
        total_credit += credit

        rows.append({
            "code": account.code,
            "account": account,
            "name": account.name,
            "type": account.get_account_type_display(),
            "debit": debit,
            "credit": credit,
            "balance": balance,
        })

    return rows, total_debit, total_credit


def build_profit_loss_standard_rows(company, date_from, date_to):
    totals = get_account_totals(company, date_from, date_to)

    rows = []

    section_totals = {
        "revenue": MONEY_ZERO,
        "cogs": MONEY_ZERO,
        "expense": MONEY_ZERO,
        "other_income": MONEY_ZERO,
        "other_expense": MONEY_ZERO,
    }

    for account in detail_accounts(company).filter(report_type=ChartOfAccount.REPORT_PROFIT_LOSS):
        row_total = totals.get(account.id, {"debit": MONEY_ZERO, "credit": MONEY_ZERO})
        amount = pl_signed_balance(account, row_total["debit"], row_total["credit"])

        if amount == 0:
            continue

        section_totals[account.account_type] += amount

        rows.append({
            "account": account,
            "account_label": f"{account.code} - {account.name}",
            "type": account.get_account_type_display(),
            "section": account.report_section,
            "amount": amount,
        })

    total_revenue = section_totals["revenue"]
    total_cogs = section_totals["cogs"]
    gross_profit = total_revenue - total_cogs
    total_expense = section_totals["expense"]
    total_other_income = section_totals["other_income"]
    total_other_expense = section_totals["other_expense"]
    net_profit = gross_profit - total_expense + total_other_income - total_other_expense

    summary = {
        "total_revenue": total_revenue,
        "total_cogs": total_cogs,
        "gross_profit": gross_profit,
        "total_expense": total_expense,
        "total_other_income": total_other_income,
        "total_other_expense": total_other_expense,
        "net_profit": net_profit,
    }

    return rows, summary


def build_balance_sheet_rows(company, date_to):
    totals = get_account_totals(company, None, date_to)

    rows = []
    total_assets = MONEY_ZERO
    total_liabilities = MONEY_ZERO
    total_equity = MONEY_ZERO

    accounts = detail_accounts(company).filter(report_type=ChartOfAccount.REPORT_BALANCE_SHEET)

    for account in accounts:
        row_total = totals.get(account.id, {"debit": MONEY_ZERO, "credit": MONEY_ZERO})
        amount = account_signed_balance(account, row_total["debit"], row_total["credit"])

        if amount == 0:
            continue

        if account.account_type == ChartOfAccount.ACCOUNT_TYPE_ASSET:
            total_assets += amount
        elif account.account_type == ChartOfAccount.ACCOUNT_TYPE_LIABILITY:
            total_liabilities += amount
        elif account.account_type == ChartOfAccount.ACCOUNT_TYPE_EQUITY:
            total_equity += amount

        rows.append({
            "account": account,
            "account_label": f"{account.code} - {account.name}",
            "type": account.get_account_type_display(),
            "section": account.report_section,
            "amount": amount,
        })

    balance_check = total_assets - total_liabilities - total_equity

    summary = {
        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "total_equity": total_equity,
        "balance_check": balance_check,
    }

    return rows, summary


# =========================================================
# 1. TRIAL BALANCE - EXCEL + PDF
# =========================================================

def export_trial_balance(company, date_from, date_to):
    wb, ws = make_workbook("Trial Balance")
    add_report_title(ws, company, "Trial Balance", date_from, date_to, 6)

    headers = ["Code", "Account", "Type", "Debit", "Credit", "Balance"]
    ws.append([])
    ws.append(headers)
    style_header(ws, 5, len(headers))

    rows, total_debit, total_credit = build_trial_balance_rows(company, date_from, date_to)

    for row in rows:
        ws.append([
            row["code"],
            row["name"],
            row["type"],
            row["debit"],
            row["credit"],
            row["balance"],
        ])

    ws.append(["", "", "TOTAL", total_debit, total_credit, total_debit - total_credit])

    style_money_columns(ws, [4, 5, 6])
    auto_width(ws)
    return write_response(wb, "Trial Balance.xlsx")


def pdf_trial_balance(company, date_from, date_to):
    rows, total_debit, total_credit = build_trial_balance_rows(company, date_from, date_to)

    pdf_rows = []

    for row in rows:
        pdf_rows.append([
            row["code"],
            row["name"],
            row["type"],
            fmt_money(row["debit"]),
            fmt_money(row["credit"]),
            fmt_money(row["balance"]),
        ])

    pdf_rows.append([
        "",
        "",
        "TOTAL",
        fmt_money(total_debit),
        fmt_money(total_credit),
        fmt_money(total_debit - total_credit),
    ])

    return simple_pdf_report(
        company,
        "Trial Balance",
        report_subtitle(date_from, date_to),
        ["Code", "Account", "Type", "Debit", "Credit", "Balance"],
        pdf_rows,
        "Trial Balance.pdf",
    )


# =========================================================
# 2. JOURNAL REPORT - EXCEL + PDF
# =========================================================

def export_journal_report(company, date_from, date_to):
    wb, ws = make_workbook("Journal Report")
    add_report_title(ws, company, "Journal Report", date_from, date_to, 8)

    headers = [
        "Date",
        "Journal No",
        "Reference No",
        "Description",
        "Account Code",
        "Account Name",
        "Debit",
        "Credit",
    ]
    ws.append([])
    ws.append(headers)
    style_header(ws, 5, len(headers))

    entries = (
        JournalEntry.objects
        .filter(company=company, status=JournalEntry.STATUS_POSTED)
        .prefetch_related("lines", "lines__account")
        .order_by("entry_date", "id")
    )

    if date_from:
        entries = entries.filter(entry_date__gte=date_from)

    if date_to:
        entries = entries.filter(entry_date__lte=date_to)

    for entry in entries:
        for line in entry.lines.all():
            ws.append([
                entry.entry_date,
                entry.entry_no,
                entry.reference_no,
                entry.description,
                line.account.code,
                line.account.name,
                money(line.debit),
                money(line.credit),
            ])

    style_money_columns(ws, [7, 8])
    auto_width(ws)
    return write_response(wb, "Journal Report.xlsx")


def pdf_journal_report(company, date_from, date_to):
    rows = []

    entries = (
        JournalEntry.objects
        .filter(company=company, status=JournalEntry.STATUS_POSTED)
        .prefetch_related("lines", "lines__account")
        .order_by("entry_date", "id")
    )

    if date_from:
        entries = entries.filter(entry_date__gte=date_from)

    if date_to:
        entries = entries.filter(entry_date__lte=date_to)

    for entry in entries:
        for line in entry.lines.all():
            rows.append([
                entry.entry_date.strftime("%d-%m-%Y"),
                entry.entry_no,
                entry.reference_no,
                line.account.code,
                line.account.name,
                fmt_money(line.debit),
                fmt_money(line.credit),
            ])

    return simple_pdf_report(
        company,
        "Journal Report",
        report_subtitle(date_from, date_to),
        ["Date", "Journal No", "Ref", "Code", "Account", "Debit", "Credit"],
        rows,
        "Journal Report.pdf",
    )


# =========================================================
# 3. GENERAL LEDGER - EXCEL + PDF
# =========================================================

def export_general_ledger(company, date_from, date_to):
    wb, ws = make_workbook("General Ledger")
    add_report_title(ws, company, "General Ledger", date_from, date_to, 9)

    headers = [
        "Account",
        "Transaction",
        "Date",
        "Reference No",
        "Name",
        "Memo",
        "Debit",
        "Credit",
        "Balance",
    ]
    ws.append([])
    ws.append(headers)
    style_header(ws, 5, len(headers))

    accounts = detail_accounts(company)

    for account in accounts:
        lines = posted_lines(company, date_from, date_to).filter(account=account).order_by(
            "journal_entry__entry_date",
            "journal_entry_id",
            "id",
        )

        if not lines.exists():
            continue

        balance = MONEY_ZERO

        ws.append([f"{account.code} - {account.name}", "", "", "", "", "", "", "", ""])

        for line in lines:
            debit = money(line.debit)
            credit = money(line.credit)

            if account.normal_balance == ChartOfAccount.NORMAL_CREDIT:
                balance += credit - debit
            else:
                balance += debit - credit

            ws.append([
                "",
                line.journal_entry.entry_no,
                line.journal_entry.entry_date,
                line.journal_entry.reference_no,
                "",
                line.description or line.journal_entry.description,
                debit,
                credit,
                balance,
            ])

    style_money_columns(ws, [7, 8, 9])
    auto_width(ws)
    return write_response(wb, "General Ledger.xlsx")


def pdf_general_ledger(company, date_from, date_to):
    rows = []

    for account in detail_accounts(company):
        lines = posted_lines(company, date_from, date_to).filter(account=account).order_by(
            "journal_entry__entry_date",
            "journal_entry_id",
            "id",
        )

        if not lines.exists():
            continue

        balance = MONEY_ZERO

        rows.append([f"{account.code} - {account.name}", "", "", "", "", "", "", ""])

        for line in lines:
            debit = money(line.debit)
            credit = money(line.credit)

            if account.normal_balance == ChartOfAccount.NORMAL_CREDIT:
                balance += credit - debit
            else:
                balance += debit - credit

            rows.append([
                "",
                line.journal_entry.entry_no,
                line.journal_entry.entry_date.strftime("%d-%m-%Y"),
                line.journal_entry.reference_no,
                line.description or line.journal_entry.description,
                fmt_money(debit),
                fmt_money(credit),
                fmt_money(balance),
            ])

    return simple_pdf_report(
        company,
        "General Ledger",
        report_subtitle(date_from, date_to),
        ["Account", "Transaction", "Date", "Ref", "Memo", "Debit", "Credit", "Balance"],
        rows,
        "General Ledger.pdf",
    )


# =========================================================
# 4. PROFIT LOSS STANDARD - EXCEL + PDF
# =========================================================

def export_profit_loss_standard(company, date_from, date_to):
    wb, ws = make_workbook("ProfitLoss Standard")
    add_report_title(ws, company, "Profit/Loss Standard", date_from, date_to, 4)

    headers = ["Account", "Type", "Section", "Amount"]
    ws.append([])
    ws.append(headers)
    style_header(ws, 5, len(headers))

    rows, summary = build_profit_loss_standard_rows(company, date_from, date_to)

    for row in rows:
        ws.append([
            row["account_label"],
            row["type"],
            row["section"],
            row["amount"],
        ])

    ws.append([])
    ws.append(["Total Revenue", "", "", summary["total_revenue"]])
    ws.append(["Total COGS", "", "", summary["total_cogs"]])
    ws.append(["Gross Profit", "", "", summary["gross_profit"]])
    ws.append(["Total Expense", "", "", summary["total_expense"]])
    ws.append(["Other Income", "", "", summary["total_other_income"]])
    ws.append(["Other Expense", "", "", summary["total_other_expense"]])
    ws.append(["Net Profit / Loss", "", "", summary["net_profit"]])

    style_money_columns(ws, [4])
    auto_width(ws)
    return write_response(wb, "ProfitLoss Standard.xlsx")


def pdf_profit_loss_standard(company, date_from, date_to):
    rows, summary = build_profit_loss_standard_rows(company, date_from, date_to)

    pdf_rows = []

    for row in rows:
        pdf_rows.append([
            row["account_label"],
            row["type"],
            row["section"],
            fmt_money(row["amount"]),
        ])

    pdf_rows += [
        ["", "", "", ""],
        ["Total Revenue", "", "", fmt_money(summary["total_revenue"])],
        ["Total COGS", "", "", fmt_money(summary["total_cogs"])],
        ["Gross Profit", "", "", fmt_money(summary["gross_profit"])],
        ["Total Expense", "", "", fmt_money(summary["total_expense"])],
        ["Other Income", "", "", fmt_money(summary["total_other_income"])],
        ["Other Expense", "", "", fmt_money(summary["total_other_expense"])],
        ["Net Profit / Loss", "", "", fmt_money(summary["net_profit"])],
    ]

    return simple_pdf_report(
        company,
        "Profit/Loss Standard",
        report_subtitle(date_from, date_to),
        ["Account", "Type", "Section", "Amount"],
        pdf_rows,
        "ProfitLoss Standard.pdf",
        landscape_page=False,
    )


# =========================================================
# 5. PROFIT LOSS BY MONTH - EXCEL + PDF
# =========================================================

def export_profit_loss_by_month(company, date_from, date_to):
    wb, ws = make_workbook("ProfitLoss by Month")

    months = list(range(1, 13))
    last_col = 14

    add_report_title(ws, company, "Profit/Loss by Month", date_from, date_to, last_col)

    headers = ["Account"] + [f"{month_abbr[m]}-{date_from.year}" for m in months] + ["Total"]
    ws.append([])
    ws.append(headers)
    style_header(ws, 5, len(headers))

    accounts = detail_accounts(company).filter(report_type=ChartOfAccount.REPORT_PROFIT_LOSS)

    for account in accounts:
        row = [f"{account.code}-{account.name}"]
        row_total = MONEY_ZERO

        for m in months:
            month_start = date(date_from.year, m, 1)

            if m == 12:
                month_end = date(date_from.year, 12, 31)
            else:
                month_end = date(date_from.year, m + 1, 1)

            if m != 12:
                qs = posted_lines(company, month_start, None).filter(
                    account=account,
                    journal_entry__entry_date__lt=month_end,
                )
            else:
                qs = posted_lines(company, month_start, month_end).filter(account=account)

            sums = qs.aggregate(debit=Sum("debit"), credit=Sum("credit"))
            amount = pl_signed_balance(account, sums["debit"], sums["credit"])

            row.append(amount)
            row_total += amount

        row.append(row_total)

        if row_total != 0:
            ws.append(row)

    style_money_columns(ws, list(range(2, 15)))
    auto_width(ws)
    return write_response(wb, "ProfitLoss by Month.xlsx")


def pdf_profit_loss_by_month(company, date_from, date_to):
    months = list(range(1, 13))
    headers = ["Account"] + [f"{month_abbr[m]}-{date_from.year}" for m in months] + ["Total"]

    pdf_rows = []

    accounts = detail_accounts(company).filter(report_type=ChartOfAccount.REPORT_PROFIT_LOSS)

    for account in accounts:
        row = [f"{account.code}-{account.name}"]
        row_total = MONEY_ZERO

        for m in months:
            month_start = date(date_from.year, m, 1)

            if m == 12:
                month_end = date(date_from.year, 12, 31)
            else:
                month_end = date(date_from.year, m + 1, 1)

            if m != 12:
                qs = posted_lines(company, month_start, None).filter(
                    account=account,
                    journal_entry__entry_date__lt=month_end,
                )
            else:
                qs = posted_lines(company, month_start, month_end).filter(account=account)

            sums = qs.aggregate(debit=Sum("debit"), credit=Sum("credit"))
            amount = pl_signed_balance(account, sums["debit"], sums["credit"])

            row.append(fmt_money(amount))
            row_total += amount

        row.append(fmt_money(row_total))

        if row_total != 0:
            pdf_rows.append(row)

    return simple_pdf_report(
        company,
        "Profit/Loss by Month",
        report_subtitle(date_from, date_to),
        headers,
        pdf_rows,
        "ProfitLoss by Month.pdf",
        landscape_page=True,
    )


# =========================================================
# 6. BALANCE SHEET - EXCEL + PDF
# =========================================================

def export_balance_sheet(company, date_from, date_to):
    wb, ws = make_workbook("Balance Sheet")
    add_report_title(ws, company, "Balance Sheet", None, date_to, 4)

    headers = ["Account", "Type", "Section", "Amount"]
    ws.append([])
    ws.append(headers)
    style_header(ws, 5, len(headers))

    rows, summary = build_balance_sheet_rows(company, date_to)

    for row in rows:
        ws.append([
            row["account_label"],
            row["type"],
            row["section"],
            row["amount"],
        ])

    ws.append([])
    ws.append(["Total Assets", "", "", summary["total_assets"]])
    ws.append(["Total Liabilities", "", "", summary["total_liabilities"]])
    ws.append(["Total Equity", "", "", summary["total_equity"]])
    ws.append(["Balance Check", "", "", summary["balance_check"]])

    style_money_columns(ws, [4])
    auto_width(ws)
    return write_response(wb, "Balance Sheet.xlsx")


def pdf_balance_sheet(company, date_from, date_to):
    rows, summary = build_balance_sheet_rows(company, date_to)

    pdf_rows = []

    for row in rows:
        pdf_rows.append([
            row["account_label"],
            row["type"],
            row["section"],
            fmt_money(row["amount"]),
        ])

    pdf_rows += [
        ["", "", "", ""],
        ["Total Assets", "", "", fmt_money(summary["total_assets"])],
        ["Total Liabilities", "", "", fmt_money(summary["total_liabilities"])],
        ["Total Equity", "", "", fmt_money(summary["total_equity"])],
        ["Balance Check", "", "", fmt_money(summary["balance_check"])],
    ]

    return simple_pdf_report(
        company,
        "Balance Sheet",
        report_subtitle(date_to=date_to, as_of=True),
        ["Account", "Type", "Section", "Amount"],
        pdf_rows,
        "Balance Sheet.pdf",
        landscape_page=False,
    )


# =========================================================
# 7. CASH FLOW - EXCEL + PDF
# =========================================================

def export_cash_flow(company, date_from, date_to):
    wb, ws = make_workbook("Cash Flow")
    add_report_title(ws, company, "Cash Flow", date_from, date_to, 4)

    headers = ["Account", "Type", "Section", "Cash Movement"]
    ws.append([])
    ws.append(headers)
    style_header(ws, 5, len(headers))

    cash_accounts = detail_accounts(company).filter(
        account_type=ChartOfAccount.ACCOUNT_TYPE_ASSET,
        name__icontains="cash",
    )

    total_cash_change = MONEY_ZERO

    for account in cash_accounts:
        sums = posted_lines(company, date_from, date_to).filter(account=account).aggregate(
            debit=Sum("debit"),
            credit=Sum("credit"),
        )
        amount = account_signed_balance(account, sums["debit"], sums["credit"])

        if amount == 0:
            continue

        total_cash_change += amount

        ws.append([
            f"{account.code} - {account.name}",
            account.get_account_type_display(),
            account.report_section,
            amount,
        ])

    ws.append([])
    ws.append(["Net Cash Increase / Decrease", "", "", total_cash_change])

    style_money_columns(ws, [4])
    auto_width(ws)
    return write_response(wb, "Cash Flow.xlsx")


def pdf_cash_flow(company, date_from, date_to):
    rows = []
    total_cash_change = MONEY_ZERO

    cash_accounts = detail_accounts(company).filter(
        account_type=ChartOfAccount.ACCOUNT_TYPE_ASSET,
        name__icontains="cash",
    )

    for account in cash_accounts:
        sums = posted_lines(company, date_from, date_to).filter(account=account).aggregate(
            debit=Sum("debit"),
            credit=Sum("credit"),
        )
        amount = account_signed_balance(account, sums["debit"], sums["credit"])

        if amount == 0:
            continue

        total_cash_change += amount

        rows.append([
            f"{account.code} - {account.name}",
            account.get_account_type_display(),
            account.report_section,
            fmt_money(amount),
        ])

    rows.append(["", "", "Net Cash Increase / Decrease", fmt_money(total_cash_change)])

    return simple_pdf_report(
        company,
        "Cash Flow",
        report_subtitle(date_from, date_to),
        ["Account", "Type", "Section", "Cash Movement"],
        rows,
        "Cash Flow.pdf",
        landscape_page=False,
    )


# =========================================================
# 8. AP AGING - EXCEL + PDF
# =========================================================

def export_ap_aging(company, date_from, date_to):
    try:
        from vendors.models import VendorTransaction
    except Exception:
        wb, ws = make_workbook("AP Aging")
        ws.append(["Vendor app not found."])
        return write_response(wb, "AP Aging Summary Detail.xlsx")

    wb, ws = make_workbook("AP Aging")
    add_report_title(ws, company, "AP Aging Summary Detail", None, date_to, 9)

    headers = ["Vendor", "Date", "Number", "Type", "Current", "1-30", "31-60", "61-90", "Over 90"]
    ws.append([])
    ws.append(headers)
    style_header(ws, 5, len(headers))

    qs = VendorTransaction.objects.filter(
        company=company,
        status=VendorTransaction.STATUS_POSTED,
        transaction_type__in=[
            VendorTransaction.TYPE_PURCHASE_ORDER,
            VendorTransaction.TYPE_ADJUSTMENT,
        ],
        transaction_date__lte=date_to,
    ).select_related("vendor")

    for tx in qs.order_by("vendor__name", "transaction_date"):
        age = (date_to - tx.transaction_date).days
        amount = money(tx.amount)

        buckets = [MONEY_ZERO, MONEY_ZERO, MONEY_ZERO, MONEY_ZERO, MONEY_ZERO]

        if age <= 0:
            buckets[0] = amount
        elif age <= 30:
            buckets[1] = amount
        elif age <= 60:
            buckets[2] = amount
        elif age <= 90:
            buckets[3] = amount
        else:
            buckets[4] = amount

        ws.append([
            tx.vendor.name if tx.vendor else "-",
            tx.transaction_date,
            tx.number,
            tx.get_transaction_type_display(),
            *buckets,
        ])

    style_money_columns(ws, [5, 6, 7, 8, 9])
    auto_width(ws)
    return write_response(wb, "AP Aging Summary Detail.xlsx")


def pdf_ap_aging(company, date_from, date_to):
    try:
        from vendors.models import VendorTransaction
    except Exception:
        return simple_pdf_report(
            company,
            "AP Aging Summary Detail",
            "",
            ["Error"],
            [["Vendor app not found"]],
            "AP Aging Summary Detail.pdf",
        )

    rows = []

    qs = VendorTransaction.objects.filter(
        company=company,
        status=VendorTransaction.STATUS_POSTED,
        transaction_type__in=[
            VendorTransaction.TYPE_PURCHASE_ORDER,
            VendorTransaction.TYPE_ADJUSTMENT,
        ],
        transaction_date__lte=date_to,
    ).select_related("vendor")

    for tx in qs.order_by("vendor__name", "transaction_date"):
        age = (date_to - tx.transaction_date).days
        amount = money(tx.amount)

        buckets = [MONEY_ZERO, MONEY_ZERO, MONEY_ZERO, MONEY_ZERO, MONEY_ZERO]

        if age <= 0:
            buckets[0] = amount
        elif age <= 30:
            buckets[1] = amount
        elif age <= 60:
            buckets[2] = amount
        elif age <= 90:
            buckets[3] = amount
        else:
            buckets[4] = amount

        rows.append([
            tx.vendor.name if tx.vendor else "-",
            tx.transaction_date.strftime("%d-%m-%Y"),
            tx.number,
            tx.get_transaction_type_display(),
            *[fmt_money(x) for x in buckets],
        ])

    return simple_pdf_report(
        company,
        "AP Aging Summary Detail",
        report_subtitle(date_to=date_to, as_of=True),
        ["Vendor", "Date", "Number", "Type", "Current", "1-30", "31-60", "61-90", "Over 90"],
        rows,
        "AP Aging Summary Detail.pdf",
    )


# =========================================================
# 9. AR AGING - EXCEL + PDF
# =========================================================

def export_ar_aging(company, date_from, date_to):
    try:
        from customers.models import CustomerTransaction
    except Exception:
        wb, ws = make_workbook("AR Aging")
        ws.append(["Customer app not found."])
        return write_response(wb, "AR Aging Summary Detail.xlsx")

    wb, ws = make_workbook("AR Aging")
    add_report_title(ws, company, "AR Aging Summary Detail", None, date_to, 9)

    headers = ["Customer", "Date", "Number", "Type", "Current", "1-30", "31-60", "61-90", "Over 90"]
    ws.append([])
    ws.append(headers)
    style_header(ws, 5, len(headers))

    qs = CustomerTransaction.objects.filter(
        company=company,
        status=CustomerTransaction.STATUS_POSTED,
        transaction_type__in=[
            CustomerTransaction.TYPE_INVOICE,
            CustomerTransaction.TYPE_ADJUSTMENT,
        ],
        transaction_date__lte=date_to,
    ).select_related("customer")

    for tx in qs.order_by("customer__name", "transaction_date"):
        age = (date_to - tx.transaction_date).days
        amount = money(tx.amount)

        buckets = [MONEY_ZERO, MONEY_ZERO, MONEY_ZERO, MONEY_ZERO, MONEY_ZERO]

        if age <= 0:
            buckets[0] = amount
        elif age <= 30:
            buckets[1] = amount
        elif age <= 60:
            buckets[2] = amount
        elif age <= 90:
            buckets[3] = amount
        else:
            buckets[4] = amount

        ws.append([
            tx.customer.name,
            tx.transaction_date,
            tx.number,
            tx.get_transaction_type_display(),
            *buckets,
        ])

    style_money_columns(ws, [5, 6, 7, 8, 9])
    auto_width(ws)
    return write_response(wb, "AR Aging Summary Detail.xlsx")


def pdf_ar_aging(company, date_from, date_to):
    try:
        from customers.models import CustomerTransaction
    except Exception:
        return simple_pdf_report(
            company,
            "AR Aging Summary Detail",
            "",
            ["Error"],
            [["Customer app not found"]],
            "AR Aging Summary Detail.pdf",
        )

    rows = []

    qs = CustomerTransaction.objects.filter(
        company=company,
        status=CustomerTransaction.STATUS_POSTED,
        transaction_type__in=[
            CustomerTransaction.TYPE_INVOICE,
            CustomerTransaction.TYPE_ADJUSTMENT,
        ],
        transaction_date__lte=date_to,
    ).select_related("customer")

    for tx in qs.order_by("customer__name", "transaction_date"):
        age = (date_to - tx.transaction_date).days
        amount = money(tx.amount)

        buckets = [MONEY_ZERO, MONEY_ZERO, MONEY_ZERO, MONEY_ZERO, MONEY_ZERO]

        if age <= 0:
            buckets[0] = amount
        elif age <= 30:
            buckets[1] = amount
        elif age <= 60:
            buckets[2] = amount
        elif age <= 90:
            buckets[3] = amount
        else:
            buckets[4] = amount

        rows.append([
            tx.customer.name,
            tx.transaction_date.strftime("%d-%m-%Y"),
            tx.number,
            tx.get_transaction_type_display(),
            *[fmt_money(x) for x in buckets],
        ])

    return simple_pdf_report(
        company,
        "AR Aging Summary Detail",
        report_subtitle(date_to=date_to, as_of=True),
        ["Customer", "Date", "Number", "Type", "Current", "1-30", "31-60", "61-90", "Over 90"],
        rows,
        "AR Aging Summary Detail.pdf",
    )


# =========================================================
# ROUTER VIEW
# =========================================================

REPORT_EXPORTS = {
    "profit-loss-by-month": export_profit_loss_by_month,
    "general-ledger": export_general_ledger,
    "ap-aging": export_ap_aging,
    "ar-aging": export_ar_aging,
    "balance-sheet": export_balance_sheet,
    "profit-loss-standard": export_profit_loss_standard,
    "journal-report": export_journal_report,
    "trial-balance": export_trial_balance,
    "cash-flow": export_cash_flow,
}


PDF_REPORT_EXPORTS = {
    "profit-loss-by-month": pdf_profit_loss_by_month,
    "general-ledger": pdf_general_ledger,
    "ap-aging": pdf_ap_aging,
    "ar-aging": pdf_ar_aging,
    "balance-sheet": pdf_balance_sheet,
    "profit-loss-standard": pdf_profit_loss_standard,
    "journal-report": pdf_journal_report,
    "trial-balance": pdf_trial_balance,
    "cash-flow": pdf_cash_flow,
}


@login_required
def export_accounting_report(request, report_slug):
    company, response = require_company_access(request)
    if response:
        return response

    date_from, date_to = get_date_range(request)
    export_format = (request.GET.get("format") or "xlsx").lower()

    if export_format == "pdf":
        pdf_exporter = PDF_REPORT_EXPORTS.get(report_slug)

        if not pdf_exporter:
            messages.error(request, "PDF report not found.")
            return redirect("accounting_reports")

        return pdf_exporter(company, date_from, date_to)

    exporter = REPORT_EXPORTS.get(report_slug)

    if not exporter:
        messages.error(request, "Report not found.")
        return redirect("accounting_reports")

    return exporter(company, date_from, date_to)