from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from accounting.models import JournalEntry, JournalEntryLine
from core.models import Company

from .forms import (
    PurchaseBillExpenseLineFormSet,
    PurchaseBillForm,
    PurchaseBillItemLineFormSet,
    VendorForm,
    VendorTransactionForm,
)
from .models import PurchaseBill, Vendor, VendorPayment, VendorTransaction


# =========================================================
# HELPERS
# =========================================================

def get_selected_company(request):
    company_id = request.session.get("selected_company_id")
    if not company_id:
        return None
    return Company.objects.filter(id=company_id, is_active=True).first()


def can_access_company(user, company):
    if not company:
        return False
    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)
    if profile and getattr(profile, "user_type", None) == "client":
        return profile.company_id == company.id

    if hasattr(company, "assigned_staff"):
        return company.assigned_staff.filter(id=user.id).exists()

    return False


def require_company_access(request):
    company = get_selected_company(request)
    if not company:
        messages.warning(request, "Please select a company first.")
        return None, redirect("company_list")

    if not can_access_company(request.user, company):
        messages.error(request, "You do not have permission to access this company.")
        return None, redirect("company_list")

    return company, None


def get_posted_status():
    return getattr(JournalEntry, "STATUS_POSTED", "posted")


def create_vendor_journal(transaction_obj, user):
    """Legacy single-line journal generator."""
    if transaction_obj.status != VendorTransaction.STATUS_POSTED:
        return None

    with transaction.atomic():
        old_entry = transaction_obj.journal_entry
        if old_entry:
            old_entry.delete()

        entry = JournalEntry.objects.create(
            company=transaction_obj.company,
            entry_date=transaction_obj.transaction_date,
            reference_no=transaction_obj.number or transaction_obj.po_number,
            description=f"{transaction_obj.get_transaction_type_display()} - {transaction_obj.vendor.name if transaction_obj.vendor else 'No Vendor'}",
            status=get_posted_status(),
            created_by=user,
        )

        JournalEntryLine.objects.create(
            journal_entry=entry,
            account=transaction_obj.debit_account,
            description=transaction_obj.memo or transaction_obj.get_transaction_type_display(),
            debit=transaction_obj.amount,
            credit=Decimal("0.00"),
        )
        JournalEntryLine.objects.create(
            journal_entry=entry,
            account=transaction_obj.credit_account,
            description=transaction_obj.memo or transaction_obj.get_transaction_type_display(),
            debit=Decimal("0.00"),
            credit=transaction_obj.amount,
        )

        transaction_obj.journal_entry = entry
        transaction_obj.save(update_fields=["journal_entry"])
        return entry


def _formset_has_rows(formset):
    for form in formset.forms:
        if not hasattr(form, "cleaned_data"):
            continue
        data = form.cleaned_data
        if data and not data.get("DELETE") and not (form.instance.pk is None and not form.has_changed()):
            return True
    return False


def validate_purchase_bill_lines(bill_form, item_formset, expense_formset):
    """Business validation that spans the header and both line formsets."""
    valid = True
    has_item = _formset_has_rows(item_formset)
    has_expense = _formset_has_rows(expense_formset)

    if not has_item and not has_expense:
        bill_form.add_error(None, "Add at least one Item/Inventory line or Expense line.")
        valid = False

    vat_total = Decimal("0.00")

    for line_form in item_formset.forms:
        data = getattr(line_form, "cleaned_data", None) or {}
        if not data or data.get("DELETE") or (line_form.instance.pk is None and not line_form.has_changed()):
            continue

        item = data.get("item")
        account = data.get("account")
        qty = data.get("qty") or Decimal("0.00")
        unit_cost = data.get("unit_cost") or Decimal("0.00")
        vat = data.get("vat_amount") or Decimal("0.00")
        vat_total += vat

        if qty <= 0:
            line_form.add_error("qty", "Quantity must be more than zero.")
            valid = False
        if unit_cost < 0:
            line_form.add_error("unit_cost", "Unit cost cannot be negative.")
            valid = False
        if item and not account and not item.inventory_account:
            line_form.add_error("account", "Select an account or set Inventory Account on this Item.")
            valid = False

    for line_form in expense_formset.forms:
        data = getattr(line_form, "cleaned_data", None) or {}
        if not data or data.get("DELETE") or (line_form.instance.pk is None and not line_form.has_changed()):
            continue

        amount = data.get("amount") or Decimal("0.00")
        vat = data.get("vat_amount") or Decimal("0.00")
        vat_total += vat
        if amount <= 0:
            line_form.add_error("amount", "Amount must be more than zero.")
            valid = False

    if vat_total > 0 and not bill_form.cleaned_data.get("input_vat_account"):
        bill_form.add_error("input_vat_account", "Select an Input VAT account because VAT was entered on a line.")
        valid = False

    return valid


def create_purchase_bill_journal(bill, user):
    """Create a multi-line journal from one Purchase Bill.

    Dr Inventory/Expense accounts (one line per purchase line)
    Dr Input VAT (if used)
    Cr Accounts Payable for grand total
    """
    if bill.status != PurchaseBill.STATUS_POSTED:
        return None

    bill.recalculate_totals(save=True)
    if bill.total_amount <= 0:
        raise ValueError("Purchase Bill total must be more than zero.")
    if bill.vat_total > 0 and not bill.input_vat_account_id:
        raise ValueError("Input VAT account is required when VAT is entered.")

    with transaction.atomic():
        if bill.journal_entry_id:
            bill.journal_entry.delete()
            bill.journal_entry = None

        entry = JournalEntry.objects.create(
            company=bill.company,
            entry_date=bill.bill_date,
            reference_no=bill.number or bill.po_number,
            description=f"Purchase Bill - {bill.vendor.name}",
            status=get_posted_status(),
            created_by=user,
        )

        for line in bill.item_lines.select_related("item", "account", "item__inventory_account"):
            account = line.account or line.item.inventory_account
            if not account:
                raise ValueError(f"Item {line.item} does not have an Inventory Account and no line account was selected.")
            JournalEntryLine.objects.create(
                journal_entry=entry,
                account=account,
                description=line.description or line.item.name,
                debit=line.line_amount,
                credit=Decimal("0.00"),
            )

        for line in bill.expense_lines.select_related("account"):
            JournalEntryLine.objects.create(
                journal_entry=entry,
                account=line.account,
                description=line.description,
                debit=line.amount,
                credit=Decimal("0.00"),
            )

        if bill.vat_total > 0:
            JournalEntryLine.objects.create(
                journal_entry=entry,
                account=bill.input_vat_account,
                description=f"Input VAT - {bill.number or bill.vendor.name}",
                debit=bill.vat_total,
                credit=Decimal("0.00"),
            )

        JournalEntryLine.objects.create(
            journal_entry=entry,
            account=bill.accounts_payable_account,
            description=f"Accounts Payable - {bill.vendor.name}",
            debit=Decimal("0.00"),
            credit=bill.total_amount,
        )

        bill.journal_entry = entry
        bill.save(update_fields=["journal_entry", "updated_at"])
        return entry


# =========================================================
# VENDOR CENTER + EXCEL
# =========================================================

@login_required
def vendor_center(request):
    company, response = require_company_access(request)
    if response:
        return response

    # Sample-style toolbar actions: bulk status, safe delete, and merge.
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        selected_ids = [int(x) for x in request.POST.getlist("selected") if str(x).isdigit()]
        selected = Vendor.objects.filter(company=company, id__in=selected_ids)

        if action in {"activate", "deactivate"}:
            if not selected_ids:
                messages.error(request, "Select at least one vendor first.")
            else:
                selected.update(is_active=(action == "activate"))
                messages.success(request, f"{selected.count()} vendor(s) updated.")
            return redirect("vendor_center")

        if action == "delete":
            if not selected_ids:
                messages.error(request, "Select at least one vendor first.")
            else:
                deleted = deactivated = 0
                for vendor in selected:
                    has_history = (
                        vendor.transactions.exists()
                        or vendor.purchase_bills.exists()
                        or vendor.payments.exists()
                    )
                    if has_history:
                        vendor.is_active = False
                        vendor.save(update_fields=["is_active"])
                        deactivated += 1
                    else:
                        vendor.delete()
                        deleted += 1
                messages.success(request, f"Deleted {deleted}; deactivated {deactivated} vendor(s) with accounting history.")
            return redirect("vendor_center")

        if action == "merge":
            target_id = request.POST.get("merge_target")
            source_ids = [x for x in selected_ids if str(x) != str(target_id)]
            target = Vendor.objects.filter(company=company, id=target_id).first()
            if not target or not source_ids:
                messages.error(request, "Choose a target vendor and at least one other vendor to merge.")
            else:
                with transaction.atomic():
                    sources = Vendor.objects.filter(company=company, id__in=source_ids)
                    for source in sources:
                        VendorTransaction.objects.filter(company=company, vendor=source).update(vendor=target)
                        PurchaseBill.objects.filter(company=company, vendor=source).update(vendor=target)
                        VendorPayment.objects.filter(company=company, vendor=source).update(vendor=target)
                        target.opening_balance = (target.opening_balance or Decimal("0")) + (source.opening_balance or Decimal("0"))
                        source.delete()
                    target.save(update_fields=["opening_balance"])
                messages.success(request, "Vendor records merged successfully; transaction history was preserved.")
            return redirect(f"{request.path}?vendor={target_id}" if target else "vendor_center")

    query = (request.GET.get("q") or "").strip()
    vendors = Vendor.objects.filter(company=company)
    if query:
        vendors = vendors.filter(
            Q(code__icontains=query) | Q(name__icontains=query) | Q(phone__icontains=query)
            | Q(email__icontains=query) | Q(contact_person__icontains=query)
        )
    vendors = vendors.order_by("name")

    total_vendors = vendors.count()
    active_vendors = vendors.filter(is_active=True).count()
    posted_transactions = VendorTransaction.objects.filter(company=company, status=VendorTransaction.STATUS_POSTED)
    legacy_purchase = posted_transactions.filter(transaction_type=VendorTransaction.TYPE_PURCHASE_ORDER).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    total_cash_expense = posted_transactions.filter(transaction_type=VendorTransaction.TYPE_CASH_EXPENSE).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    total_payment = posted_transactions.filter(transaction_type=VendorTransaction.TYPE_VENDOR_PAYMENT).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    new_purchase = PurchaseBill.objects.filter(company=company, status=PurchaseBill.STATUS_POSTED).aggregate(total=Sum("total_amount"))["total"] or Decimal("0.00")
    opening_total = Vendor.objects.filter(company=company).aggregate(total=Sum("opening_balance"))["total"] or Decimal("0.00")
    total_purchase = legacy_purchase + new_purchase
    ap_balance = opening_total + total_purchase - total_payment

    selected_vendor = None
    selected_vendor_id = request.GET.get("vendor")
    if selected_vendor_id:
        selected_vendor = Vendor.objects.filter(company=company, id=selected_vendor_id).first()
    if selected_vendor is None:
        selected_vendor = vendors.first()

    recent_activity = []
    if selected_vendor:
        for bill in selected_vendor.purchase_bills.filter(company=company).order_by("-bill_date", "-id")[:30]:
            recent_activity.append({"date": bill.bill_date, "type": "Bill", "number": bill.number or f"BILL-{bill.id}", "amount": bill.total_amount, "status": bill.get_status_display(), "memo": bill.memo, "url_name": "purchase_bill_detail", "pk": bill.pk})
        for pay in selected_vendor.payments.filter(company=company).order_by("-payment_date", "-id")[:30]:
            recent_activity.append({"date": pay.payment_date, "type": "Payment", "number": pay.number or f"PAY-{pay.id}", "amount": -pay.total_amount, "status": pay.get_status_display(), "memo": pay.memo, "url_name": "", "pk": pay.pk})
        for txn in selected_vendor.transactions.filter(company=company).order_by("-transaction_date", "-id")[:30]:
            recent_activity.append({"date": txn.transaction_date, "type": txn.get_transaction_type_display(), "number": txn.number or f"TXN-{txn.id}", "amount": txn.credit_amount - txn.debit_amount, "status": txn.get_status_display(), "memo": txn.memo, "url_name": "vendor_transaction_detail", "pk": txn.pk})
        recent_activity.sort(key=lambda x: (x["date"], x["pk"]), reverse=True)
        recent_activity = recent_activity[:50]

    return render(request, "vendors/vendor_center.html", {
        "company": company, "vendors": vendors, "query": query,
        "total_vendors": total_vendors, "active_vendors": active_vendors,
        "total_purchase": total_purchase, "total_cash_expense": total_cash_expense,
        "total_payment": total_payment, "ap_balance": ap_balance,
        "selected_vendor": selected_vendor, "recent_activity": recent_activity,
    })


@login_required
def vendor_export_excel(request):
    company, response = require_company_access(request)
    if response:
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendors"
    headers = ["Vendor Code", "Vendor Name", "Phone", "Email", "Contact Person", "Address", "Opening Balance", "Memo", "Active"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for vendor in Vendor.objects.filter(company=company).order_by("name"):
        ws.append([
            vendor.code,
            vendor.name,
            vendor.phone,
            vendor.email,
            vendor.contact_person,
            vendor.address,
            float(vendor.opening_balance or 0),
            vendor.memo,
            "Yes" if vendor.is_active else "No",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    safe_company = "".join(c if c.isalnum() or c in "-_" else "_" for c in company.name)
    response["Content-Disposition"] = f'attachment; filename="{safe_company}_Vendor_List.xlsx"'
    return response


@login_required
def vendor_import_excel(request):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "Please choose an Excel file.")
        else:
            try:
                wb = load_workbook(excel_file, data_only=True)
                ws = wb.active
                headers = {str(cell.value or "").strip().lower(): idx for idx, cell in enumerate(ws[1])}

                def col(*names):
                    for name in names:
                        if name.lower() in headers:
                            return headers[name.lower()]
                    return None

                code_col = col("vendor code", "code", "vendor id")
                name_col = col("vendor name", "name")
                if name_col is None:
                    raise ValueError("Excel must contain a 'Vendor Name' column.")

                created_count = 0
                updated_count = 0
                skipped_count = 0

                with transaction.atomic():
                    for values in ws.iter_rows(min_row=2, values_only=True):
                        name = str(values[name_col] or "").strip()
                        if not name:
                            skipped_count += 1
                            continue
                        code = str(values[code_col] or "").strip() if code_col is not None else ""

                        lookup = {"company": company, "code": code} if code else {"company": company, "name": name}
                        vendor = Vendor.objects.filter(**lookup).first()
                        created = vendor is None
                        if created:
                            vendor = Vendor(company=company, created_by=request.user)

                        vendor.code = code
                        vendor.name = name

                        mapping = {
                            "phone": col("phone"),
                            "email": col("email"),
                            "contact_person": col("contact person", "contact"),
                            "address": col("address"),
                            "memo": col("memo", "note"),
                        }
                        for field, index in mapping.items():
                            if index is not None:
                                setattr(vendor, field, str(values[index] or "").strip())

                        opening_col = col("opening balance", "opening_balance")
                        if opening_col is not None and values[opening_col] not in (None, ""):
                            vendor.opening_balance = Decimal(str(values[opening_col]))

                        active_col = col("active", "is active", "status")
                        if active_col is not None:
                            raw = str(values[active_col] or "").strip().lower()
                            vendor.is_active = raw not in {"no", "false", "0", "inactive"}

                        vendor.full_clean()
                        vendor.save()
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1

                messages.success(
                    request,
                    f"Vendor import completed: {created_count} created, {updated_count} updated, {skipped_count} skipped.",
                )
                return redirect("vendor_center")
            except Exception as exc:
                messages.error(request, f"Import failed: {exc}")

    return render(request, "vendors/vendor_import.html", {"company": company})


@login_required
def vendor_create(request):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method == "POST":
        form = VendorForm(request.POST, company=company)
        if form.is_valid():
            vendor = form.save(commit=False)
            vendor.company = company
            vendor.created_by = request.user
            vendor.save()
            messages.success(request, f"Vendor {vendor.name} created successfully.")
            return redirect("vendor_center")
    else:
        form = VendorForm(company=company, initial={"is_active": True})

    return render(request, "vendors/vendor_form.html", {
        "company": company,
        "form": form,
        "page_title": "Create Vendor",
        "button_text": "Create Vendor",
    })


@login_required
def vendor_edit(request, vendor_id):
    company, response = require_company_access(request)
    if response:
        return response

    vendor = get_object_or_404(Vendor, id=vendor_id, company=company)
    if request.method == "POST":
        form = VendorForm(request.POST, instance=vendor, company=company)
        if form.is_valid():
            form.save()
            messages.success(request, f"Vendor {vendor.name} updated successfully.")
            return redirect("vendor_center")
    else:
        form = VendorForm(instance=vendor, company=company)

    return render(request, "vendors/vendor_form.html", {
        "company": company,
        "form": form,
        "vendor": vendor,
        "page_title": "Edit Vendor",
        "button_text": "Save Changes",
    })


# =========================================================
# PURCHASE BILLS (NEW MULTI-LINE WORKFLOW)
# =========================================================

@login_required
def purchase_bill_list(request):
    company, response = require_company_access(request)
    if response:
        return response

    q = (request.GET.get("q") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    bills = PurchaseBill.objects.filter(company=company).select_related("vendor", "accounts_payable_account", "journal_entry")
    if q:
        bills = bills.filter(
            Q(vendor__name__icontains=q)
            | Q(vendor__code__icontains=q)
            | Q(number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(memo__icontains=q)
        )
    if date_from:
        bills = bills.filter(bill_date__gte=date_from)
    if date_to:
        bills = bills.filter(bill_date__lte=date_to)

    bills = bills.order_by("-bill_date", "-id")
    total_amount = bills.aggregate(total=Sum("total_amount"))["total"] or Decimal("0.00")

    return render(request, "vendors/purchase_bill_list.html", {
        "company": company,
        "bills": bills,
        "query": q,
        "date_from": date_from,
        "date_to": date_to,
        "total_amount": total_amount,
    })


def _purchase_bill_formsets(request, company, instance):
    kwargs = {
        "instance": instance,
        "prefix": "items",
        "form_kwargs": {"company": company},
    }
    expense_kwargs = {
        "instance": instance,
        "prefix": "expenses",
        "form_kwargs": {"company": company},
    }
    if request.method == "POST":
        kwargs["data"] = request.POST
        expense_kwargs["data"] = request.POST
    return PurchaseBillItemLineFormSet(**kwargs), PurchaseBillExpenseLineFormSet(**expense_kwargs)


@login_required
def purchase_bill_create(request):
    company, response = require_company_access(request)
    if response:
        return response

    bill = PurchaseBill(company=company, created_by=request.user, status=PurchaseBill.STATUS_POSTED)
    if request.method == "POST":
        form = PurchaseBillForm(request.POST, instance=bill, company=company)
    else:
        form = PurchaseBillForm(instance=bill, company=company, initial={"bill_date": timezone.localdate()})

    item_formset, expense_formset = _purchase_bill_formsets(request, company, bill)

    if request.method == "POST" and form.is_valid() and item_formset.is_valid() and expense_formset.is_valid():
        if validate_purchase_bill_lines(form, item_formset, expense_formset):
            try:
                with transaction.atomic():
                    bill = form.save(commit=False)
                    bill.company = company
                    bill.created_by = request.user
                    bill.status = PurchaseBill.STATUS_POSTED
                    bill.save()

                    item_formset.instance = bill
                    item_formset.save()
                    expense_formset.instance = bill
                    expense_formset.save()

                    bill.recalculate_totals(save=True)
                    create_purchase_bill_journal(bill, request.user)

                messages.success(request, f"Purchase Bill {bill.number or bill.id} saved and journal generated.")
                if request.POST.get("save_action") == "save_new":
                    return redirect("purchase_bill_create")
                return redirect("purchase_bill_list")
            except Exception as exc:
                messages.error(request, f"Could not save Purchase Bill: {exc}")

    return render(request, "vendors/purchase_bill_form.html", {
        "company": company,
        "form": form,
        "item_formset": item_formset,
        "expense_formset": expense_formset,
        "page_title": "Purchase Order / Bill",
        "is_edit": False,
    })


@login_required
def purchase_bill_edit(request, bill_id):
    company, response = require_company_access(request)
    if response:
        return response

    bill = get_object_or_404(PurchaseBill, id=bill_id, company=company)
    form = PurchaseBillForm(request.POST or None, instance=bill, company=company)
    item_formset, expense_formset = _purchase_bill_formsets(request, company, bill)

    if request.method == "POST" and form.is_valid() and item_formset.is_valid() and expense_formset.is_valid():
        if validate_purchase_bill_lines(form, item_formset, expense_formset):
            try:
                with transaction.atomic():
                    bill = form.save(commit=False)
                    bill.company = company
                    bill.status = PurchaseBill.STATUS_POSTED
                    bill.save()
                    item_formset.save()
                    expense_formset.save()
                    bill.recalculate_totals(save=True)
                    create_purchase_bill_journal(bill, request.user)

                messages.success(request, "Purchase Bill updated and journal regenerated.")
                if request.POST.get("save_action") == "save_new":
                    return redirect("purchase_bill_create")
                return redirect("purchase_bill_list")
            except Exception as exc:
                messages.error(request, f"Could not update Purchase Bill: {exc}")

    return render(request, "vendors/purchase_bill_form.html", {
        "company": company,
        "form": form,
        "item_formset": item_formset,
        "expense_formset": expense_formset,
        "page_title": "Edit Purchase Bill",
        "is_edit": True,
        "bill": bill,
    })


@login_required
def purchase_bill_detail(request, bill_id):
    company, response = require_company_access(request)
    if response:
        return response

    bill = get_object_or_404(
        PurchaseBill.objects.select_related(
            "vendor",
            "accounts_payable_account",
            "input_vat_account",
            "journal_entry",
        ).prefetch_related("item_lines__item", "expense_lines__account"),
        id=bill_id,
        company=company,
    )
    return render(request, "vendors/purchase_bill_detail.html", {"company": company, "bill": bill})


# =========================================================
# LEGACY VENDOR TRANSACTIONS (PAYMENT/CASH EXPENSE FOR NOW)
# =========================================================

@login_required
def vendor_transaction_list(request):
    company, response = require_company_access(request)
    if response:
        return response

    query = (request.GET.get("q") or "").strip()
    tran_type = (request.GET.get("type") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    if not date_from and not date_to:
        today = timezone.localdate()
        date_from = today.replace(day=1).strftime("%Y-%m-%d")
        date_to = today.strftime("%Y-%m-%d")

    transactions = VendorTransaction.objects.filter(company=company)
    if query:
        transactions = transactions.filter(
            Q(vendor__name__icontains=query)
            | Q(vendor__code__icontains=query)
            | Q(number__icontains=query)
            | Q(po_number__icontains=query)
            | Q(memo__icontains=query)
        )
    if tran_type:
        transactions = transactions.filter(transaction_type=tran_type)
    if date_from:
        transactions = transactions.filter(transaction_date__gte=date_from)
    if date_to:
        transactions = transactions.filter(transaction_date__lte=date_to)

    transactions = transactions.select_related("vendor", "debit_account", "credit_account", "journal_entry").order_by("-transaction_date", "-id")
    total_amount = transactions.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

    return render(request, "vendors/vendor_transaction_list.html", {
        "company": company,
        "transactions": transactions,
        "query": query,
        "tran_type": tran_type,
        "date_from": date_from,
        "date_to": date_to,
        "total_amount": total_amount,
        "type_choices": VendorTransaction.TYPE_CHOICES,
    })


@login_required
def vendor_transaction_create(request, transaction_type=None):
    company, response = require_company_access(request)
    if response:
        return response

    if transaction_type == VendorTransaction.TYPE_PURCHASE_ORDER:
        return redirect("purchase_bill_create")

    if request.method == "POST":
        form = VendorTransactionForm(request.POST, company=company, transaction_type=transaction_type)
        if form.is_valid():
            with transaction.atomic():
                vendor_transaction = form.save(commit=False)
                vendor_transaction.company = company
                vendor_transaction.created_by = request.user
                if transaction_type:
                    vendor_transaction.transaction_type = transaction_type
                vendor_transaction.save()
                create_vendor_journal(vendor_transaction, request.user)
            messages.success(request, "Vendor transaction saved and journal entry generated.")
            return redirect("vendor_transaction_list")
    else:
        form = VendorTransactionForm(
            company=company,
            transaction_type=transaction_type,
            initial={
                "transaction_date": timezone.localdate(),
                "status": VendorTransaction.STATUS_POSTED,
                "transaction_type": transaction_type or VendorTransaction.TYPE_CASH_EXPENSE,
            },
        )

    title_map = {
        VendorTransaction.TYPE_CASH_EXPENSE: "Cash Expense",
        VendorTransaction.TYPE_VENDOR_PAYMENT: "Vendor Payment",
    }
    return render(request, "vendors/vendor_transaction_form.html", {
        "company": company,
        "form": form,
        "page_title": title_map.get(transaction_type, "Vendor Transaction"),
        "button_text": "Save & Generate Journal",
    })


@login_required
def vendor_transaction_detail(request, transaction_id):
    company, response = require_company_access(request)
    if response:
        return response

    vendor_transaction = get_object_or_404(
        VendorTransaction.objects.select_related("vendor", "debit_account", "credit_account", "journal_entry"),
        id=transaction_id,
        company=company,
    )
    return render(request, "vendors/vendor_transaction_detail.html", {
        "company": company,
        "transaction": vendor_transaction,
    })

# =========================================================
# CUSTOMER REQUIREMENT PAYMENT IMPLEMENTATION
# =========================================================
from .forms import VendorPaymentForm, VendorPaymentAllocationFormSet, VendorPaymentOtherChargeFormSet
from .models import VendorPayment, VendorPaymentAllocation, VendorPaymentOtherCharge


def create_vendor_payment_journal(payment,user):
    if payment.status!=VendorPayment.STATUS_POSTED or payment.total_amount<=0:return None
    with transaction.atomic():
        if payment.journal_entry_id:payment.journal_entry.delete()
        entry=JournalEntry.objects.create(company=payment.company,entry_date=payment.payment_date,reference_no=payment.number,description=f"Payment - {payment.vendor.name}",status=get_posted_status(),created_by=user)
        ap_total=payment.allocations.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        if ap_total>0:JournalEntryLine.objects.create(journal_entry=entry,account=payment.accounts_payable_account,description="Apply to bills",debit=ap_total,credit=Decimal("0.00"))
        for ch in payment.other_charges.select_related("account"):
            if ch.amount>0:JournalEntryLine.objects.create(journal_entry=entry,account=ch.account,description=ch.memo,debit=ch.amount,credit=Decimal("0.00"))
        JournalEntryLine.objects.create(journal_entry=entry,account=payment.payment_account,description=payment.memo or "Vendor payment",debit=Decimal("0.00"),credit=payment.total_amount)
        payment.journal_entry=entry;payment.save(update_fields=["journal_entry"])
    return entry


def _vendor_payment_formset(formset_class,data,instance,company,prefix,extra=None):
    return formset_class(data=data,instance=instance,prefix=prefix,form_kwargs={"company":company,**(extra or {})}) if data is not None else formset_class(instance=instance,prefix=prefix,form_kwargs={"company":company,**(extra or {})})

@login_required
def vendor_payment_new(request):
    company,response=require_company_access(request)
    if response:return response
    payment=VendorPayment(company=company,created_by=request.user)
    vendor_id=(request.POST.get("vendor") if request.method=="POST" else request.GET.get("vendor")) or None
    form=VendorPaymentForm(request.POST or None,instance=payment,company=company)
    allocations=_vendor_payment_formset(VendorPaymentAllocationFormSet,request.POST if request.method=="POST" else None,payment,company,"alloc",{"vendor_id":vendor_id})
    charges=_vendor_payment_formset(VendorPaymentOtherChargeFormSet,request.POST if request.method=="POST" else None,payment,company,"charge")
    if request.method=="POST" and form.is_valid() and allocations.is_valid() and charges.is_valid():
        try:
            with transaction.atomic():
                payment=form.save(commit=False);payment.company=company;payment.created_by=request.user;payment.status=VendorPayment.STATUS_POSTED;payment.save()
                allocations.instance=payment;allocations.save();charges.instance=payment;charges.save();payment.recalculate_total();create_vendor_payment_journal(payment,request.user)
            messages.success(request,"Payment saved successfully.")
            if request.POST.get("save_action")=="save_new":return redirect("vendor_payment_create")
            return redirect("vendor_center")
        except Exception as exc:messages.error(request,f"Could not save payment: {exc}")
    bills=PurchaseBill.objects.filter(company=company,status=PurchaseBill.STATUS_POSTED)
    if vendor_id:bills=bills.filter(vendor_id=vendor_id)
    return render(request,"vendors/payment_form.html",{"company":company,"form":form,"allocation_formset":allocations,"charge_formset":charges,"bills":bills})
