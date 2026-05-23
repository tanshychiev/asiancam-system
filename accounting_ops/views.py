from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from accounting.models import ChartOfAccount, JournalEntry, JournalEntryLine
from core.models import Company

from .forms import (
    BankAccountForm,
    BankDepositForm,
    BankReconcileForm,
    BankRuleForm,
    ImportUploadForm,
    LandedCostAllocationForm,
)
from .importers import process_import
from .models import (
    BankAccount,
    BankDeposit,
    BankReconcile,
    BankRule,
    ImportHistory,
    LandedCostAllocation,
)


# =========================================================
# COMPANY ACCESS
# =========================================================

def get_selected_company(request):
    company_id = request.session.get("selected_company_id")

    if not company_id:
        return None

    return Company.objects.filter(
        id=company_id,
        is_active=True,
    ).first()


def can_access_company(user, company):
    if not company:
        return False

    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)

    if profile and getattr(profile, "user_type", None) == "client":
        return profile.company_id == company.id

    if hasattr(company, "assigned_staff"):
        return company.assigned_staff.filter(id=user.id).exists()

    return False


def require_company_access(request):
    company = get_selected_company(request)

    if not company:
        messages.warning(request, "Please select a company first.")
        return None, redirect("company_list")

    if not can_access_company(request.user, company):
        messages.error(request, "You do not have permission to access this company.")
        return None, redirect("company_list")

    return company, None


# =========================================================
# JOURNAL HELPERS
# =========================================================

def get_posted_status():
    return getattr(JournalEntry, "STATUS_POSTED", "posted")


def to_decimal(value):
    if value in [None, ""]:
        return Decimal("0.00")

    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0.00")


def create_simple_journal(company, user, entry_date, reference_no, description, debit_account, credit_account, amount):
    if not debit_account or not credit_account:
        return None

    if amount <= 0:
        return None

    entry = JournalEntry.objects.create(
        company=company,
        entry_date=entry_date,
        reference_no=reference_no,
        description=description,
        status=get_posted_status(),
        created_by=user,
    )

    JournalEntryLine.objects.create(
        journal_entry=entry,
        account=debit_account,
        description=description,
        debit=amount,
        credit=Decimal("0.00"),
    )

    JournalEntryLine.objects.create(
        journal_entry=entry,
        account=credit_account,
        description=description,
        debit=Decimal("0.00"),
        credit=amount,
    )

    return entry


def create_bank_deposit_journal(deposit, user):
    if deposit.status != BankDeposit.STATUS_POSTED:
        return None

    with transaction.atomic():
        old_entry = deposit.journal_entry

        if old_entry:
            old_entry.delete()

        entry = create_simple_journal(
            company=deposit.company,
            user=user,
            entry_date=deposit.deposit_date,
            reference_no=deposit.number,
            description=f"Bank Deposit - {deposit.deposit_to.name}",
            debit_account=deposit.deposit_to.chart_account,
            credit_account=deposit.source_account,
            amount=deposit.amount,
        )

        deposit.journal_entry = entry
        deposit.save(update_fields=["journal_entry"])

        return entry


def create_landed_cost_journal(allocation, user):
    if allocation.status != LandedCostAllocation.STATUS_POSTED:
        return None

    with transaction.atomic():
        old_entry = allocation.journal_entry

        if old_entry:
            old_entry.delete()

        entry = create_simple_journal(
            company=allocation.company,
            user=user,
            entry_date=allocation.allocation_date,
            reference_no=allocation.number,
            description=f"Landed Cost Allocation - {allocation.bill_no}",
            debit_account=allocation.landed_cost_account,
            credit_account=allocation.clearing_account,
            amount=allocation.amount,
        )

        allocation.journal_entry = entry
        allocation.save(update_fields=["journal_entry"])

        return entry


# =========================================================
# BANK ACCOUNT
# =========================================================

@login_required
def bank_account_list(request):
    company, response = require_company_access(request)
    if response:
        return response

    query = (request.GET.get("q") or "").strip()

    rows = BankAccount.objects.filter(company=company).select_related("chart_account")

    if query:
        rows = rows.filter(
            Q(name__icontains=query)
            | Q(bank_name__icontains=query)
            | Q(account_number__icontains=query)
            | Q(chart_account__name__icontains=query)
            | Q(chart_account__code__icontains=query)
        )

    rows = rows.order_by("name")

    return render(request, "accounting_ops/bank_account_list.html", {
        "company": company,
        "rows": rows,
        "query": query,
    })


@login_required
def bank_account_create(request):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method == "POST":
        form = BankAccountForm(request.POST, company=company)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = company
            obj.save()

            messages.success(request, "Bank account saved successfully.")
            return redirect("ops_bank_account_list")
    else:
        form = BankAccountForm(
            company=company,
            initial={"is_active": True},
        )

    return render(request, "accounting_ops/form.html", {
        "company": company,
        "form": form,
        "page_title": "Create Bank Account",
        "button_text": "Save",
    })


# =========================================================
# BANK DEPOSIT
# =========================================================

@login_required
def bank_deposit_list(request):
    company, response = require_company_access(request)
    if response:
        return response

    query = (request.GET.get("q") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    if not date_from and not date_to:
        today = timezone.localdate()
        date_from = today.replace(day=1).strftime("%Y-%m-%d")
        date_to = today.strftime("%Y-%m-%d")

    rows = BankDeposit.objects.filter(company=company).select_related(
        "deposit_to",
        "source_account",
        "journal_entry",
    )

    if query:
        rows = rows.filter(
            Q(number__icontains=query)
            | Q(memo__icontains=query)
            | Q(deposit_to__name__icontains=query)
        )

    if date_from:
        rows = rows.filter(deposit_date__gte=date_from)

    if date_to:
        rows = rows.filter(deposit_date__lte=date_to)

    rows = rows.order_by("-deposit_date", "-id")

    return render(request, "accounting_ops/bank_deposit_list.html", {
        "company": company,
        "rows": rows,
        "query": query,
        "date_from": date_from,
        "date_to": date_to,
    })


@login_required
def bank_deposit_create(request):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method == "POST":
        form = BankDepositForm(request.POST, company=company)

        if form.is_valid():
            with transaction.atomic():
                obj = form.save(commit=False)
                obj.company = company
                obj.created_by = request.user
                obj.save()
                create_bank_deposit_journal(obj, request.user)

            messages.success(request, "Bank deposit saved and journal generated.")
            return redirect("ops_bank_deposit_list")
    else:
        form = BankDepositForm(
            company=company,
            initial={
                "deposit_date": timezone.localdate(),
                "status": BankDeposit.STATUS_POSTED,
            },
        )

    return render(request, "accounting_ops/form.html", {
        "company": company,
        "form": form,
        "page_title": "Create Bank Deposit",
        "button_text": "Save & Generate Journal",
    })


# =========================================================
# BANK RULE
# =========================================================

@login_required
def bank_rule_list(request):
    company, response = require_company_access(request)
    if response:
        return response

    query = (request.GET.get("q") or "").strip()

    rows = BankRule.objects.filter(company=company).select_related("target_account")

    if query:
        rows = rows.filter(
            Q(name__icontains=query)
            | Q(keyword__icontains=query)
            | Q(memo__icontains=query)
        )

    rows = rows.order_by("priority_order", "name")

    return render(request, "accounting_ops/bank_rule_list.html", {
        "company": company,
        "rows": rows,
        "query": query,
    })


@login_required
def bank_rule_create(request):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method == "POST":
        form = BankRuleForm(request.POST, company=company)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = company
            obj.save()

            messages.success(request, "Bank rule saved successfully.")
            return redirect("ops_bank_rule_list")
    else:
        form = BankRuleForm(
            company=company,
            initial={"is_active": True},
        )

    return render(request, "accounting_ops/form.html", {
        "company": company,
        "form": form,
        "page_title": "Create Bank Rule",
        "button_text": "Save",
    })


# =========================================================
# BANK RECONCILE
# =========================================================

@login_required
def bank_reconcile_list(request):
    company, response = require_company_access(request)
    if response:
        return response

    rows = BankReconcile.objects.filter(company=company).select_related("bank_account")
    rows = rows.order_by("-reconcile_date", "-id")

    return render(request, "accounting_ops/bank_reconcile_list.html", {
        "company": company,
        "rows": rows,
    })


@login_required
def bank_reconcile_create(request):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method == "POST":
        form = BankReconcileForm(request.POST, company=company)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = company
            obj.created_by = request.user
            obj.save()

            messages.success(request, "Bank reconcile saved successfully.")
            return redirect("ops_bank_reconcile_list")
    else:
        form = BankReconcileForm(
            company=company,
            initial={"reconcile_date": timezone.localdate()},
        )

    return render(request, "accounting_ops/form.html", {
        "company": company,
        "form": form,
        "page_title": "Create Bank Reconcile",
        "button_text": "Save",
    })


# =========================================================
# LANDED COST
# =========================================================

@login_required
def landed_cost_list(request):
    company, response = require_company_access(request)
    if response:
        return response

    rows = LandedCostAllocation.objects.filter(company=company).select_related(
        "landed_cost_account",
        "clearing_account",
        "journal_entry",
    )
    rows = rows.order_by("-allocation_date", "-id")

    return render(request, "accounting_ops/landed_cost_list.html", {
        "company": company,
        "rows": rows,
    })


@login_required
def landed_cost_create(request):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method == "POST":
        form = LandedCostAllocationForm(request.POST, company=company)

        if form.is_valid():
            with transaction.atomic():
                obj = form.save(commit=False)
                obj.company = company
                obj.created_by = request.user
                obj.save()
                create_landed_cost_journal(obj, request.user)

            messages.success(request, "Landed cost saved and journal generated.")
            return redirect("ops_landed_cost_list")
    else:
        form = LandedCostAllocationForm(
            company=company,
            initial={
                "allocation_date": timezone.localdate(),
                "status": LandedCostAllocation.STATUS_POSTED,
            },
        )

    return render(request, "accounting_ops/form.html", {
        "company": company,
        "form": form,
        "page_title": "Create Landed Cost Allocation",
        "button_text": "Save & Generate Journal",
    })


# =========================================================
# FIND TRANSACTION / BATCH
# =========================================================

@login_required
def find_transaction(request):
    company, response = require_company_access(request)
    if response:
        return response

    query = (request.GET.get("q") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    entries = JournalEntry.objects.filter(company=company)

    if query:
        entries = entries.filter(
            Q(entry_no__icontains=query)
            | Q(reference_no__icontains=query)
            | Q(description__icontains=query)
            | Q(lines__account__code__icontains=query)
            | Q(lines__account__name__icontains=query)
        ).distinct()

    if date_from:
        entries = entries.filter(entry_date__gte=date_from)

    if date_to:
        entries = entries.filter(entry_date__lte=date_to)

    entries = (
        entries
        .prefetch_related("lines", "lines__account")
        .order_by("-entry_date", "-id")[:200]
    )

    return render(request, "accounting_ops/find_transaction.html", {
        "company": company,
        "query": query,
        "date_from": date_from,
        "date_to": date_to,
        "entries": entries,
    })


@login_required
def batch_transaction(request):
    company, response = require_company_access(request)
    if response:
        return response

    return render(request, "accounting_ops/batch_transaction.html", {
        "company": company,
    })


# =========================================================
# IMPORT
# =========================================================

@login_required
def import_page(request, import_type):
    company, response = require_company_access(request)
    if response:
        return response

    form = ImportUploadForm()

    if request.method == "POST":
        form = ImportUploadForm(request.POST, request.FILES)

        if form.is_valid():
            upload_file = form.cleaned_data["file"]

            if not upload_file.name.lower().endswith(".xlsx"):
                messages.error(request, "Wrong file type. Please upload .xlsx file only.")
                return redirect(request.path)

            try:
                success_rows, total_rows, errors = process_import(
                    company=company,
                    user=request.user,
                    import_type=import_type,
                    excel_file=upload_file,
                )

                upload_file.seek(0)

                ImportHistory.objects.create(
                    company=company,
                    import_type=import_type,
                    file=upload_file,
                    total_rows=total_rows,
                    success_rows=success_rows,
                    error_rows=len(errors),
                    created_by=request.user,
                    note="\n".join(errors[:30]) if errors else "Import completed successfully.",
                )

                if errors:
                    for error in errors[:10]:
                        messages.error(request, error)

                    messages.warning(
                        request,
                        f"Import finished with errors. Success: {success_rows}, Error: {len(errors)}.",
                    )
                else:
                    messages.success(
                        request,
                        f"Import completed successfully. {success_rows} rows imported.",
                    )

                return redirect(request.path)

            except Exception as e:
                messages.error(request, f"Import failed: {e}")
                return redirect(request.path)

    title_map = {
        ImportHistory.TYPE_STOCK_BALANCE: "Import Stock Balance",
        ImportHistory.TYPE_ITEM: "Import Item",
        ImportHistory.TYPE_VENDOR: "Import Vendor",
        ImportHistory.TYPE_CUSTOMER: "Import Customer",
        ImportHistory.TYPE_COA: "Import Chart of Account",
        ImportHistory.TYPE_BATCH_TRANSACTION: "Batch Transaction Import",

        ImportHistory.TYPE_TRIAL_BALANCE: "Import Trial Balance Full",
        ImportHistory.TYPE_JOURNAL_OPENING: "Import Journal Opening Balance",
        ImportHistory.TYPE_OUTSTANDING_AP: "Import Outstanding AP",
        ImportHistory.TYPE_OUTSTANDING_AR: "Import Outstanding AR",
    }

    histories = ImportHistory.objects.filter(
        company=company,
        import_type=import_type,
    )[:20]

    return render(request, "accounting_ops/import_page.html", {
        "company": company,
        "form": form,
        "import_type": import_type,
        "page_title": title_map.get(import_type, "Import"),
        "histories": histories,
    })


# =========================================================
# DOWNLOAD SAMPLE EXCEL
# =========================================================

def style_sample_sheet(ws):
    blue_fill = PatternFill("solid", fgColor="0070C0")
    white_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        length = 12

        for cell in column_cells:
            if cell.value:
                length = max(length, len(str(cell.value)) + 2)

        ws.column_dimensions[column_cells[0].column_letter].width = min(length, 35)


@login_required
def download_sample(request, import_type):
    company, response = require_company_access(request)
    if response:
        return response

    wb = Workbook()
    ws = wb.active

    filename = "sample.xlsx"

    if import_type == ImportHistory.TYPE_STOCK_BALANCE:
        filename = "Item_Open_Balance_Template.xlsx"
        ws.title = "Stock Balance"
        headers = ["* ITEM", "* BASE_QTY", "* AMOUNT", "MEMO"]
        rows = [
            ["fanta", 100, 100, "open stock"],
            ["cocacola", 100, 150, "open stock"],
            ["Olate", 100, 100, "open stock"],
            ["Bacas", 100, 120, "open stock"],
            ["Anchor", 100, 150, "open stock"],
            ["Angkor", 100, 200, "open stock"],
        ]

    elif import_type == ImportHistory.TYPE_ITEM:
        filename = "Item_Template.xlsx"
        ws.title = "Item"
        headers = [
            "ITEM_CODE",
            "* ITEM_NAME",
            "LOCAL_NAME",
            "* ITEM_TYPE",
            "* ITEM_GROUP",
            "ITEM_BRAND",
            "DESCRIPTION",
            "NEGATIVE_SALE",
            "* ALARM_QTY",
            "SEPARATE_REVENUE_EXPENSE_ACCOUNT",
            "FOR_PURCHASE",
            "FOR_SALE",
            "IS_HAVE_UNITSET",
            "* UNITSET_NAME",
            "* BASE_UNIT_1",
            "* BASE_PRICE_1",
            "BASE_BARCODE_1",
            "UNIT_2",
            "MULTIPLIER_2",
            "PRICE_2",
            "BARCODE_2",
            "UNIT_3",
            "MULTIPLIER_3",
            "PRICE_3",
            "BARCODE_3",
            "* COST",
            "ACCOUNT_ASSET_CODE",
            "ACCOUNT_COGS_CODE",
            "ACCOUNT_REVENUE_CODE",
            "MEMO",
            "DETAIL_MEMO",
        ]
        rows = [
            ["F-001", "fanta", "", "STOCK_PART", "ទឹកក្រូច", "", "", "FALSE", 0, "", "TRUE", "TRUE", "TRUE", "case/2L", "can", 1.00, "", "12can", 12, 12.00, "", "កេស", 24, 24.00, "", 0.6, "131000", "510000", "410000", "", ""],
            ["Co-001", "cocacola", "", "STOCK_PART", "ទឹកក្រូច", "", "", "FALSE", 0, "", "TRUE", "TRUE", "TRUE", "case/2L", "can", 1.50, "", "12can", 12, 18.00, "", "កេស", 24, 36.00, "", 1, "131000", "510000", "410000", "", ""],
            ["OL-001", "Olate", "", "STOCK_PART", "ទឹកក្រូច", "", "", "FALSE", 0, "", "TRUE", "TRUE", "TRUE", "case/2L", "can", 1.00, "", "12can", 12, 12.00, "", "កេស", 24, 24.00, "", 0.5, "131000", "510000", "410000", "", ""],
        ]

    elif import_type == ImportHistory.TYPE_VENDOR:
        filename = "Vendor_Template.xlsx"
        ws.title = "Vendor"
        headers = [
            "VENDOR_CODE",
            "* VENDOR_NAME",
            "LOCAL_VENDOR_NAME",
            "EMAIL",
            "PHONE",
            "CURRENCY",
            "VATTIN",
            "HOUSE_NO",
            "STREET",
            "COMMUNE",
            "DISTRICT",
            "CITY",
            "LOCAL_HOUSE_NO",
            "LOCAL_STREET",
            "LOCAL_COMMUNE",
            "LOCAL_DISTRICT",
            "LOCAL_CITY",
            "ADDRESS",
            "MEMO",
            "BRANCHES",
        ]
        rows = [
            ["V1", "Vendor 1", "អ្នកផ្គត់ផ្គង់ទី១", "", "", "Dollar USD", "K001-123456789", "E1122", "#2004", "Commune", "District", "City", "House Number", "#2004", "Khmer Commune", "Khmer District", "Phnom Penh", "Phnom Penh", "memo", "PP ; KD"],
            ["V2", "Vendor 2", "អ្នកផ្គត់ផ្គង់ទី២", "", "", "Dollar USD", "K001-123456790", "E1122", "#2004", "Commune", "District", "City", "House Number", "#2004", "Khmer Commune", "Khmer District", "Phnom Penh", "Phnom Penh", "memo", "PP"],
        ]

    elif import_type == ImportHistory.TYPE_CUSTOMER:
        filename = "Customer_Template.xlsx"
        ws.title = "Customer"
        headers = [
            "CUSTOMER_CODE",
            "* CUSTOMER_NAME",
            "SUB_OF",
            "LOCAL_CUSTOMER_NAME",
            "CUSTOMER_TYPE",
            "SALE_PERSON",
            "REGION",
            "GRADE",
            "CURRENCY",
            "PRICE_LEVEL",
            "INVOICE_TYPE",
            "IS_ALLOW_OVER_CREDIT",
            "CREDIT_LIMIT",
            "CREDIT_TERM",
            "EMAIL",
            "PHONE",
            "VATTIN",
            "HOUSE_NO",
            "STREET",
            "COMMUNE",
            "DISTRICT",
            "CITY",
            "LOCAL_HOUSE_NO",
            "LOCAL_STREET",
            "LOCAL_COMMUNE",
            "LOCAL_DISTRICT",
            "LOCAL_CITY",
            "ADDRESS",
            "MEMO",
            "CONTACT_NAME",
            "CONTACT_PHONE",
            "CONTACT_EMAIL",
            "BRANCHES",
        ]
        rows = [
            ["C1", "Customer 1", "", "អតិថិន១", "", "", "phnom penh", "", "Dollar USD", "", "TAX_INVOICE", "TRUE", 0, 0, "", "", "K001-123456789", "E1122", "#2004", "Commune", "District", "City", "House Number", "#2004", "Khmer Commune", "Khmer District", "Phnom Penh", "", "", "", "", "", "PP"],
            ["C2", "Customer 2", "", "អតិថិន២", "", "", "phnom penh", "", "Riel", "", "TAX_INVOICE", "TRUE", 0, 0, "", "", "K001-123456790", "E1123", "#2005", "Commune", "District", "City", "House Number", "#2005", "Khmer Commune", "Khmer District", "Phnom Penh", "", "", "", "", "", "KD"],
            ["C3", "Customer 3", "", "អតិថិន៣", "", "", "phnom penh", "", "Dollar USD", "", "TAX_INVOICE", "TRUE", 200, 0, "", "", "K001-123456791", "E1124", "#2006", "Commune", "District", "City", "House Number", "#2006", "Khmer Commune", "Khmer District", "Phnom Penh", "", "", "", "", "", ""],
        ]

    elif import_type == ImportHistory.TYPE_COA:
        filename = "Chart_Of_Account_Template.xlsx"
        ws.title = "Chart Of Account"
        headers = [
            "ACCOUNT_CODE",
            "* ACCOUNT_NAME",
            "LOCAL_ACCOUNT_NAME",
            "* ACCOUNT_TYPE",
            "SUB_OF",
            "DESCRIPTION",
            "ACTIVE",
        ]
        rows = [
            ["111000", "Cash and Cash Equivalents", "សាច់ប្រាក់ និងសាច់ប្រាក់សមមូល", "Bank", "", "", "TRUE"],
            ["111101", "Cash in Bank", "សាច់ប្រាក់នៅធនាគារ", "Bank", "Cash and Cash Equivalents", "Cash deposit in bank and Expense for big Amount", "TRUE"],
            ["111500", "Cash on Hand", "សាច់ប្រាក់នៅក្នុងបេឡា", "Bank", "Cash and Cash Equivalents", "for office pay in small amount", "TRUE"],
            ["120000", "Accounts Receivable", "គណនីត្រូវទទួល ឬ អតិថិជនជំពាក់", "Account Receivable", "", "Receivable due from customers", "TRUE"],
            ["131000", "Inventory Asset", "សន្និធិ", "Inventory", "", "Inventory purchase", "TRUE"],
            ["200000", "Accounts Payable", "គណនីត្រូវសង ឬ ជំពាក់អ្នកផ្គត់ផ្គង់", "Account Payable", "", "Payable to suppliers", "TRUE"],
            ["399999", "Opening Balance Equity", "បើកសមតុល្យដើមគ្រា", "Equity", "", "Opening Balance", "TRUE"],
            ["410000", "Sale Revenues", "ចំណូលពីការលក់", "Income", "", "Revenue from sales", "TRUE"],
            ["510000", "Cost Of Sales", "ថ្លៃដើមនៃការលក់", "Cost of Goods Sold", "", "COGS", "TRUE"],
        ]

    elif import_type == ImportHistory.TYPE_TRIAL_BALANCE:
        filename = "Batch_Trial_Balance_Full_Template.xlsx"
        ws.title = "Trial Balance Full"
        headers = [
            "* DATE",
            "BRANCH_CODE",
            "CLASS",
            "NUMBER",
            "MEMO",
            "ADJUST_ENTRY",
            "MEMO_DETAIL",
            "* DEBIT",
            "* CREDIT",
            "* ACCOUNT_CODE",
            "NAME",
            "BRANCH_CODE_DETAIL",
            "CLASS_DETAIL",
        ]
        rows = [
            ["31/12/2023", "", "", "OPEN-2023", "Opening Balance", "TRUE", "Opening Balance - Cash in Bank", 5760.59, "-", "111000", "", "", ""],
            ["", "", "", "", "", "", "Opening Balance - Accounts Receivable", 35392.50, "-", "150100", "", "", ""],
            ["", "", "", "", "", "", "Opening Balance - Accounts Payable", "-", 24948.00, "290100", "", "", ""],
            ["", "", "", "", "", "", "Opening Balance Equity", "-", 16205.09, "399999", "", "", ""],
        ]

    elif import_type == ImportHistory.TYPE_JOURNAL_OPENING:
        filename = "Journal_Opening_Balance_Template.xlsx"
        ws.title = "Journal Opening"
        headers = [
            "* DATE",
            "BRANCH_CODE",
            "CLASS",
            "NUMBER",
            "MEMO",
            "ADJUST_ENTRY",
            "MEMO_DETAIL",
            "* DEBIT",
            "* CREDIT",
            "* ACCOUNT_CODE",
            "NAME",
            "BRANCH_CODE_DETAIL",
            "CLASS_DETAIL",
        ]
        rows = [
            ["12/31/2023", "", "", "JV-23-001", "Opening Balance Clearing", "TRUE", "Accounts Receivable - Old", "-", 35392.50, "150100", "", "", ""],
            ["", "", "", "", "", "", "Football Inventory", "-", 730, "131100", "", "", ""],
            ["", "", "", "", "", "", "Accounts Payable - Old", 24948.00, "-", "290100", "", "", ""],
            ["", "", "", "", "", "", "Opening balance", 11174.50, "-", "399999", "", "", ""],
        ]

    elif import_type == ImportHistory.TYPE_OUTSTANDING_AP:
        filename = "Batch_Outstanding_AP_Template.xlsx"
        ws.title = "Outstanding AP"
        headers = [
            "* TRAN_TYPE",
            "* VENDOR",
            "BRANCH_CODE",
            "LOCATION",
            "* DATE",
            "BILL_NO",
            "DISCOUNT_ACCOUNT_CODE",
            "CLASS",
            "EXCHANGE_RATE",
            "DISPLAY_TAX_EXCHANGE_RATE",
            "MEMO",
            "ITEM_CODE",
            "* ITEM_NAME",
            "DESCRIPTION",
            "* UNIT_NAME",
            "* QTY",
            "COST",
            "AMOUNT",
            "DISCOUNT_AMOUNT",
            "VAT",
            "WHT",
            "GROSS_NET",
            "BRANCH_CODE_DETAIL",
            "CLASS_ITEM",
            "MEMO_EXPENSE",
            "AMOUNT_EXPENSE",
            "ACCOUNT_CODE_EXPENSE",
            "VAT_EXPENSE",
            "WHT_EXPENSE",
            "GROSS_NET_EXPENSE",
            "CUSTOMER_PROJECT_EXPENSE",
            "BRANCH_CODE_EXPENSE",
            "CLASS_EXPENSE",
        ]
        rows = [
            ["PURCHASE", "MPT Co.,", "", "", "10/02/2023", "Inv# M306", "", "", 4000, "", "Opening AP Balance", "Co-001", "Opening Balance", "", "N/A", 1, 4455, 4455, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["PURCHASE_RETURN", "Sam Shop", "", "", "31/03/2023", "Inv# M513", "", "", 4000, "", "Opening AP Balance", "", "Opening Balance", "", "N/A", 1, 1782, 1782, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        ]

    elif import_type == ImportHistory.TYPE_OUTSTANDING_AR:
        filename = "Batch_Outstanding_AR_Template.xlsx"
        ws.title = "Outstanding AR"
        headers = [
            "* TRAN_TYPE",
            "* INVOICE_TYPE",
            "BRANCH_CODE",
            "CLASS",
            "LOCATION",
            "SALE_PERSON",
            "TRUCK_NO",
            "* DATE",
            "DUE_DATE",
            "NUMBER",
            "* CUSTOMER",
            "DISCOUNT_ACCOUNT_CODE",
            "EXCHANGE_RATE",
            "DISPLAY_TAX_EXCHANGE_RATE",
            "MEMO",
            "ITEM_CODE",
            "* ITEM_NAME",
            "DESCRIPTION",
            "* UNIT_NAME",
            "* QTY",
            "PRICE",
            "AMOUNT",
            "VAT",
            "DISCOUNT_AMOUNT",
            "CLASS_ITEM",
        ]
        rows = [
            ["INVOICE", "COMMERCIAL_INVOICE", "", "", "", "", "", "23/03/2023", "23/03/2023", "Inv #SS23-034", "APK", "", 4000, 4000, "Opening AR Balance", "", "Opening Balance", "", "N/A", 1, 1815, 1815, "", "", ""],
            ["INVOICE", "COMMERCIAL_INVOICE", "", "", "", "", "", "06/03/2023", "06/03/2023", "Inv #SS23-026", "Cams Sport", "", 4000, 4000, "Opening AR Balance", "", "Opening Balance", "", "N/A", 1, 1815, 1815, "", "", ""],
        ]

    else:
        filename = "Batch_Transaction_Template.xlsx"
        ws.title = "Batch Transaction"
        headers = [
            "DATE",
            "NUMBER",
            "DEBIT_ACCOUNT_CODE",
            "CREDIT_ACCOUNT_CODE",
            "AMOUNT",
            "MEMO",
        ]
        rows = [
            ["2026-05-16", "JV-001", "650100", "111500", 100, "Rental expense"],
            ["2026-05-16", "JV-002", "650200", "111500", 50, "Utility expense"],
        ]

    ws.append(headers)

    for row in rows:
        ws.append(row)

    style_sample_sheet(ws)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

