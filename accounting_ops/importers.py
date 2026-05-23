from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from accounting.models import ChartOfAccount, JournalEntry, JournalEntryLine


# =========================================================
# BASIC HELPERS
# =========================================================

def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_header(value):
    value = clean_text(value)
    value = value.replace("*", "")
    value = value.replace(" ", "_")
    return value.upper()


def to_decimal(value):
    if value in [None, "", "-"]:
        return Decimal("0.00")

    text = str(value).replace("$", "").replace(",", "").strip()

    if text in ["", "-"]:
        return Decimal("0.00")

    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0.00")


def to_bool(value):
    text = clean_text(value).lower()

    if text in ["true", "yes", "1", "y", "active"]:
        return True

    if text in ["false", "no", "0", "n", "inactive"]:
        return False

    return False


def read_excel_rows(excel_file):
    """
    Read Excel rows and auto-detect the header row.

    This supports normal templates where row 1 is the header, and client files
    where the first row is a title like "Batch_outstanding_ap" and the real
    header starts a few rows below.
    """
    from openpyxl import load_workbook

    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return [], ["Excel file is empty."]

    header_row_index = None
    headers = []

    header_markers = {
        "ACCOUNT_CODE",
        "ITEM_CODE",
        "VENDOR",
        "VENDOR_NAME",
        "CUSTOMER",
        "CUSTOMER_NAME",
        "TRAN_TYPE",
        "DATE",
        "DEBIT_ACCOUNT_CODE",
        "CREDIT_ACCOUNT_CODE",
        "DEBIT",
        "CREDIT",
        "AMOUNT",
    }

    for index, row in enumerate(rows[:30]):
        possible_headers = [clean_header(h) for h in row]
        non_empty_headers = [h for h in possible_headers if h]

        if not non_empty_headers:
            continue

        score = len(set(non_empty_headers) & header_markers)

        if score >= 2 or "ACCOUNT_CODE" in non_empty_headers or "TRAN_TYPE" in non_empty_headers:
            header_row_index = index
            headers = possible_headers
            break

    if header_row_index is None:
        header_row_index = 0
        headers = [clean_header(h) for h in rows[0]]

    clean_rows = []

    for index, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
        data = {}
        empty = True

        for col_index, header in enumerate(headers):
            if not header:
                continue

            value = row[col_index] if col_index < len(row) else None

            if value not in [None, ""]:
                empty = False

            data[header] = value

        if not empty:
            data["_ROW"] = index
            clean_rows.append(data)

    return clean_rows, []


def get_value(row, *names):
    for name in names:
        key = clean_header(name)
        if key in row:
            return row.get(key)
    return None


def model_has_field(model_class, field_name):
    try:
        model_class._meta.get_field(field_name)
        return True
    except Exception:
        return False


def set_if_field(obj, field_name, value):
    if model_has_field(obj.__class__, field_name):
        setattr(obj, field_name, value)


def get_posted_status():
    return getattr(JournalEntry, "STATUS_POSTED", "posted")


def create_simple_journal(company, user, entry_date, reference_no, description, debit_account, credit_account, amount):
    if not debit_account or not credit_account:
        return None

    if amount <= 0:
        return None

    entry = JournalEntry.objects.create(
        company=company,
        entry_date=entry_date,
        reference_no=reference_no,
        description=description,
        status=get_posted_status(),
        created_by=user,
    )

    JournalEntryLine.objects.create(
        journal_entry=entry,
        account=debit_account,
        description=description,
        debit=amount,
        credit=Decimal("0.00"),
    )

    JournalEntryLine.objects.create(
        journal_entry=entry,
        account=credit_account,
        description=description,
        debit=Decimal("0.00"),
        credit=amount,
    )

    return entry


# =========================================================
# CHART OF ACCOUNT IMPORT
# =========================================================

def map_account_type(raw_type):
    text = clean_text(raw_type).lower()

    mapping = {
        "bank": "asset",
        "cash": "asset",
        "account receivable": "asset",
        "accounts receivable": "asset",
        "inventory": "asset",
        "advance to supplier": "asset",
        "other current asset": "asset",
        "fixed asset": "asset",

        "account payable": "liability",
        "accounts payable": "liability",
        "customer deposit": "liability",
        "other current liability": "liability",
        "liability": "liability",

        "equity": "equity",

        "income": "revenue",
        "revenue": "revenue",

        "cost of goods sold": "cogs",
        "cost of sales": "cogs",
        "cogs": "cogs",

        "operating expense": "expense",
        "expense": "expense",
    }

    return mapping.get(text, "expense")


def get_report_type(account_type):
    if account_type in ["revenue", "cogs", "expense", "other_income", "other_expense"]:
        return getattr(ChartOfAccount, "REPORT_PROFIT_LOSS", "profit_loss")

    return getattr(ChartOfAccount, "REPORT_BALANCE_SHEET", "balance_sheet")


def get_normal_balance(account_type):
    if account_type in ["asset", "expense", "cogs", "other_expense"]:
        return "debit"
    return "credit"


@transaction.atomic
def import_chart_of_accounts(company, user, excel_file):
    rows, errors = read_excel_rows(excel_file)

    success = 0

    if errors:
        return success, len(rows), errors

    created_accounts = {}

    for row in rows:
        row_number = row["_ROW"]

        code = clean_text(get_value(row, "ACCOUNT_CODE"))
        name = clean_text(get_value(row, "ACCOUNT_NAME"))

        if not code:
            errors.append(f"Row {row_number}: ACCOUNT_CODE is required.")
            continue

        if not name:
            errors.append(f"Row {row_number}: ACCOUNT_NAME is required.")
            continue

        account_type = map_account_type(get_value(row, "ACCOUNT_TYPE"))
        report_type = get_report_type(account_type)
        normal_balance = get_normal_balance(account_type)
        active = to_bool(get_value(row, "ACTIVE"))

        account, created = ChartOfAccount.objects.update_or_create(
            company=company,
            code=code,
            defaults={
                "name": name,
                "local_name": clean_text(get_value(row, "LOCAL_ACCOUNT_NAME")),
                "account_type": account_type,
                "report_type": report_type,
                "normal_balance": normal_balance,
                "description": clean_text(get_value(row, "DESCRIPTION")),
                "is_active": active,
            },
        )

        created_accounts[code] = account
        created_accounts[name.lower()] = account
        success += 1

    # second pass for parent / sub_of
    for row in rows:
        code = clean_text(get_value(row, "ACCOUNT_CODE"))
        sub_of = clean_text(get_value(row, "SUB_OF"))

        if not code or not sub_of:
            continue

        account = created_accounts.get(code)
        parent = created_accounts.get(sub_of) or created_accounts.get(sub_of.lower())

        if account and parent and account.id != parent.id:
            account.parent = parent
            account.save(update_fields=["parent"])

    return success, len(rows), errors


# =========================================================
# ITEM IMPORT
# =========================================================

@transaction.atomic
def import_items(company, user, excel_file):
    rows, errors = read_excel_rows(excel_file)
    success = 0

    if errors:
        return success, len(rows), errors

    try:
        from stock.models import Item, ItemGroup, ItemBrand, UnitSet
    except Exception as e:
        return 0, len(rows), [f"Stock app import error: {e}"]

    for row in rows:
        row_number = row["_ROW"]

        code = clean_text(get_value(row, "ITEM_CODE"))
        name = clean_text(get_value(row, "ITEM_NAME"))

        if not code:
            errors.append(f"Row {row_number}: ITEM_CODE is required.")
            continue

        if not name:
            errors.append(f"Row {row_number}: ITEM_NAME is required.")
            continue

        group_name = clean_text(get_value(row, "ITEM_GROUP"))
        brand_name = clean_text(get_value(row, "ITEM_BRAND"))
        unitset_name = clean_text(get_value(row, "UNITSET_NAME"))

        group = None
        brand = None
        unit_set = None

        if group_name:
            group, _ = ItemGroup.objects.get_or_create(
                company=company,
                name=group_name,
                defaults={"is_active": True},
            )

        if brand_name:
            brand, _ = ItemBrand.objects.get_or_create(
                company=company,
                name=brand_name,
                defaults={"is_active": True},
            )

        if unitset_name:
            unit_set, _ = UnitSet.objects.get_or_create(
                company=company,
                name=unitset_name,
                defaults={
                    "base_unit": clean_text(get_value(row, "BASE_UNIT_1")) or "pcs",
                    "default_purchase": clean_text(get_value(row, "BASE_UNIT_1")) or "pcs",
                    "default_sale": clean_text(get_value(row, "BASE_UNIT_1")) or "pcs",
                    "is_active": True,
                },
            )

        asset_account = ChartOfAccount.objects.filter(
            company=company,
            code=clean_text(get_value(row, "ACCOUNT_ASSET_CODE")),
        ).first()

        cogs_account = ChartOfAccount.objects.filter(
            company=company,
            code=clean_text(get_value(row, "ACCOUNT_COGS_CODE")),
        ).first()

        revenue_account = ChartOfAccount.objects.filter(
            company=company,
            code=clean_text(get_value(row, "ACCOUNT_REVENUE_CODE")),
        ).first()

        item_type = clean_text(get_value(row, "ITEM_TYPE")) or "STOCK_PART"
        item_type = item_type.lower()

        if item_type == "stock_part":
            item_type = getattr(Item, "TYPE_STOCK_PART", "stock_part")
        elif item_type == "stock_assemble":
            item_type = getattr(Item, "TYPE_STOCK_ASSEMBLE", "stock_assemble")
        elif item_type == "service":
            item_type = getattr(Item, "TYPE_SERVICE", "service")
        else:
            item_type = getattr(Item, "TYPE_STOCK_PART", "stock_part")

        item, created = Item.objects.update_or_create(
            company=company,
            code=code,
            defaults={
                "name": name,
                "item_type": item_type,
                "item_group": group,
                "item_brand": brand,
                "unit_set": unit_set,
                "cost_price": to_decimal(get_value(row, "COST")),
                "sale_price": to_decimal(get_value(row, "BASE_PRICE_1")),
                "inventory_account": asset_account,
                "cogs_account": cogs_account,
                "revenue_account": revenue_account,
                "memo": clean_text(get_value(row, "MEMO")),
                "is_active": True,
            },
        )

        set_if_field(item, "local_name", clean_text(get_value(row, "LOCAL_NAME")))
        set_if_field(item, "description", clean_text(get_value(row, "DESCRIPTION")))
        item.save()

        success += 1

    return success, len(rows), errors


# =========================================================
# VENDOR IMPORT
# =========================================================

@transaction.atomic
def import_vendors(company, user, excel_file):
    rows, errors = read_excel_rows(excel_file)
    success = 0

    if errors:
        return success, len(rows), errors

    try:
        from vendors.models import Vendor
    except Exception as e:
        return 0, len(rows), [f"Vendors app import error: {e}"]

    for row in rows:
        row_number = row["_ROW"]

        code = clean_text(get_value(row, "VENDOR_CODE"))
        name = clean_text(get_value(row, "VENDOR_NAME"))

        if not name:
            errors.append(f"Row {row_number}: VENDOR_NAME is required.")
            continue

        lookup = {"company": company, "name": name}

        if code and model_has_field(Vendor, "code"):
            lookup = {"company": company, "code": code}

        vendor, created = Vendor.objects.get_or_create(
            **lookup,
            defaults={"name": name},
        )

        set_if_field(vendor, "name", name)
        set_if_field(vendor, "code", code)
        set_if_field(vendor, "vendor_code", code)
        set_if_field(vendor, "local_name", clean_text(get_value(row, "LOCAL_VENDOR_NAME")))
        set_if_field(vendor, "local_vendor_name", clean_text(get_value(row, "LOCAL_VENDOR_NAME")))
        set_if_field(vendor, "email", clean_text(get_value(row, "EMAIL")))
        set_if_field(vendor, "phone", clean_text(get_value(row, "PHONE")))
        set_if_field(vendor, "currency", clean_text(get_value(row, "CURRENCY")))
        set_if_field(vendor, "vat_tin", clean_text(get_value(row, "VATTIN")))
        set_if_field(vendor, "vattin", clean_text(get_value(row, "VATTIN")))
        set_if_field(vendor, "address", clean_text(get_value(row, "ADDRESS")))
        set_if_field(vendor, "memo", clean_text(get_value(row, "MEMO")))
        set_if_field(vendor, "is_active", True)

        if model_has_field(Vendor, "created_by"):
            vendor.created_by = user

        vendor.save()
        success += 1

    return success, len(rows), errors


# =========================================================
# CUSTOMER IMPORT
# =========================================================

@transaction.atomic
def import_customers(company, user, excel_file):
    rows, errors = read_excel_rows(excel_file)
    success = 0

    if errors:
        return success, len(rows), errors

    try:
        from customers.models import Customer, CustomerType, Salesperson, PriceLevel, Region
    except Exception as e:
        return 0, len(rows), [f"Customers app import error: {e}"]

    for row in rows:
        row_number = row["_ROW"]

        code = clean_text(get_value(row, "CUSTOMER_CODE"))
        name = clean_text(get_value(row, "CUSTOMER_NAME"))

        if not name:
            errors.append(f"Row {row_number}: CUSTOMER_NAME is required.")
            continue

        customer_type = None
        salesperson = None
        price_level = None
        region = None

        customer_type_name = clean_text(get_value(row, "CUSTOMER_TYPE"))
        salesperson_name = clean_text(get_value(row, "SALE_PERSON"))
        price_level_name = clean_text(get_value(row, "PRICE_LEVEL"))
        region_name = clean_text(get_value(row, "REGION"))

        if customer_type_name:
            customer_type, _ = CustomerType.objects.get_or_create(
                company=company,
                name=customer_type_name,
                defaults={"is_active": True},
            )

        if salesperson_name:
            salesperson, _ = Salesperson.objects.get_or_create(
                company=company,
                name=salesperson_name,
                defaults={"is_active": True},
            )

        if price_level_name:
            price_level, _ = PriceLevel.objects.get_or_create(
                company=company,
                name=price_level_name,
                defaults={"is_active": True},
            )

        if region_name:
            region, _ = Region.objects.get_or_create(
                company=company,
                name=region_name,
                defaults={"is_active": True},
            )

        customer, created = Customer.objects.update_or_create(
            company=company,
            name=name,
            defaults={
                "code": code,
                "local_name": clean_text(get_value(row, "LOCAL_CUSTOMER_NAME")),
                "phone": clean_text(get_value(row, "PHONE")),
                "email": clean_text(get_value(row, "EMAIL")),
                "address": clean_text(get_value(row, "ADDRESS")),
                "customer_type": customer_type,
                "salesperson": salesperson,
                "price_level": price_level,
                "region": region,
                "opening_balance": to_decimal(get_value(row, "CREDIT_LIMIT")),
                "memo": clean_text(get_value(row, "MEMO")),
                "is_active": True,
            },
        )

        if model_has_field(Customer, "created_by") and created:
            customer.created_by = user
            customer.save(update_fields=["created_by"])

        success += 1

    return success, len(rows), errors


# =========================================================
# STOCK BALANCE IMPORT
# =========================================================

@transaction.atomic
def import_stock_balance(company, user, excel_file):
    rows, errors = read_excel_rows(excel_file)
    success = 0

    if errors:
        return success, len(rows), errors

    try:
        from stock.models import Item, StockDocument, StockDocumentLine, Warehouse
    except Exception as e:
        return 0, len(rows), [f"Stock app import error: {e}"]

    inventory_account = ChartOfAccount.objects.filter(company=company, code="131000").first()
    equity_account = ChartOfAccount.objects.filter(company=company, code="399999").first()

    if not inventory_account:
        errors.append("Missing account 131000 Inventory Asset. Please import Chart of Account first.")

    if not equity_account:
        errors.append("Missing account 399999 Opening Balance Equity. Please import Chart of Account first.")

    warehouse = Warehouse.objects.filter(company=company, is_active=True).first()

    if not warehouse:
        warehouse = Warehouse.objects.create(
            company=company,
            name="Main Warehouse",
            code="MAIN",
            is_active=True,
        )

    if errors:
        return success, len(rows), errors

    doc = StockDocument.objects.create(
        company=company,
        document_type=StockDocument.TYPE_ADJUSTMENT,
        number=f"OPEN-STOCK-{timezone.localdate().strftime('%Y%m%d')}",
        document_date=timezone.localdate(),
        warehouse=warehouse,
        debit_account=inventory_account,
        credit_account=equity_account,
        memo="Opening stock balance import",
        status=StockDocument.STATUS_POSTED,
        created_by=user,
    )

    total_amount = Decimal("0.00")

    for row in rows:
        row_number = row["_ROW"]

        item_name = clean_text(get_value(row, "ITEM"))
        qty = to_decimal(get_value(row, "BASE_QTY"))
        amount = to_decimal(get_value(row, "AMOUNT"))
        memo = clean_text(get_value(row, "MEMO"))

        if not item_name:
            errors.append(f"Row {row_number}: ITEM is required.")
            continue

        if qty <= 0:
            errors.append(f"Row {row_number}: BASE_QTY must be more than zero.")
            continue

        item = Item.objects.filter(company=company, name=item_name).first()

        if not item:
            item = Item.objects.create(
                company=company,
                code=item_name.upper().replace(" ", "-")[:50],
                name=item_name,
                item_type=getattr(Item, "TYPE_STOCK_PART", "stock_part"),
                cost_price=(amount / qty) if qty else Decimal("0.00"),
                sale_price=Decimal("0.00"),
                inventory_account=inventory_account,
                is_active=True,
            )

        unit_cost = amount / qty if qty else Decimal("0.00")

        StockDocumentLine.objects.create(
            document=doc,
            item=item,
            qty=qty,
            unit_cost=unit_cost,
            memo=memo,
        )

        total_amount += amount
        success += 1

    if total_amount > 0:
        entry = create_simple_journal(
            company=company,
            user=user,
            entry_date=doc.document_date,
            reference_no=doc.number,
            description="Opening stock balance import",
            debit_account=inventory_account,
            credit_account=equity_account,
            amount=total_amount,
        )

        doc.journal_entry = entry
        doc.save(update_fields=["journal_entry"])

    return success, len(rows), errors


# =========================================================
# BATCH TRANSACTION IMPORT
# =========================================================

@transaction.atomic
def import_batch_transaction(company, user, excel_file):
    rows, errors = read_excel_rows(excel_file)
    success = 0

    if errors:
        return success, len(rows), errors

    for row in rows:
        row_number = row["_ROW"]

        date_value = get_value(row, "DATE", "TRAN_DATE")
        reference_no = clean_text(get_value(row, "NO", "NUMBER", "REFERENCE"))
        memo = clean_text(get_value(row, "MEMO"))
        debit_code = clean_text(get_value(row, "DEBIT_ACCOUNT_CODE", "DR_ACCOUNT_CODE"))
        credit_code = clean_text(get_value(row, "CREDIT_ACCOUNT_CODE", "CR_ACCOUNT_CODE"))
        amount = to_decimal(get_value(row, "AMOUNT"))

        if not debit_code:
            errors.append(f"Row {row_number}: DEBIT_ACCOUNT_CODE is required.")
            continue

        if not credit_code:
            errors.append(f"Row {row_number}: CREDIT_ACCOUNT_CODE is required.")
            continue

        if amount <= 0:
            errors.append(f"Row {row_number}: AMOUNT must be more than zero.")
            continue

        debit_account = ChartOfAccount.objects.filter(company=company, code=debit_code).first()
        credit_account = ChartOfAccount.objects.filter(company=company, code=credit_code).first()

        if not debit_account:
            errors.append(f"Row {row_number}: Debit account {debit_code} not found.")
            continue

        if not credit_account:
            errors.append(f"Row {row_number}: Credit account {credit_code} not found.")
            continue

        entry_date = timezone.localdate()

        create_simple_journal(
            company=company,
            user=user,
            entry_date=entry_date,
            reference_no=reference_no,
            description=memo or "Batch transaction import",
            debit_account=debit_account,
            credit_account=credit_account,
            amount=amount,
        )

        success += 1

    return success, len(rows), errors



# =========================================================
# OPENING BALANCE / CLIENT SAMPLE IMPORTS
# =========================================================

def parse_excel_date(value):
    if not value:
        return timezone.localdate()

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = clean_text(value)

    for fmt in ["%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"]:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return timezone.localdate()


def get_account_by_code_or_name(company, codes=None, names=None):
    codes = [code for code in (codes or []) if code]
    names = [name for name in (names or []) if name]

    if codes:
        account = ChartOfAccount.objects.filter(
            company=company,
            code__in=codes,
            is_active=True,
            is_group=False,
        ).first()

        if account:
            return account

    if names:
        query = Q()
        for name in names:
            query |= Q(name__icontains=name)

        return ChartOfAccount.objects.filter(
            company=company,
            is_active=True,
            is_group=False,
        ).filter(query).first()

    return None


@transaction.atomic
def import_journal_style_opening(company, user, excel_file, default_description="Opening Balance"):
    """
    Supports:
    - Batch_Trial_Balance_Full.xlsx
    - Journal Entry batch_Clearing Opening Balance.xlsx

    Expected columns:
    DATE, NUMBER, MEMO, MEMO_DETAIL, DEBIT, CREDIT, ACCOUNT_CODE
    Blank DATE/NUMBER/MEMO rows will reuse the previous non-empty value.
    """
    rows, errors = read_excel_rows(excel_file)
    success = 0

    if errors:
        return success, len(rows), errors

    grouped = {}
    last_date = None
    last_number = ""
    last_memo = ""

    for row in rows:
        row_number = row["_ROW"]

        date_value = get_value(row, "DATE")
        number = clean_text(get_value(row, "NUMBER", "NO", "REFERENCE_NO"))
        memo = clean_text(get_value(row, "MEMO", "ENTRY_DESCRIPTION"))
        memo_detail = clean_text(get_value(row, "MEMO_DETAIL", "LINE_DESCRIPTION", "NAME", "DESCRIPTION"))
        account_code = clean_text(get_value(row, "ACCOUNT_CODE"))
        debit = to_decimal(get_value(row, "DEBIT"))
        credit = to_decimal(get_value(row, "CREDIT"))

        if date_value:
            last_date = parse_excel_date(date_value)

        if number:
            last_number = number

        if memo:
            last_memo = memo

        entry_date = last_date or timezone.localdate()
        reference_no = last_number or default_description
        entry_description = last_memo or default_description

        if not account_code and debit == 0 and credit == 0:
            continue

        if not account_code:
            errors.append(f"Row {row_number}: ACCOUNT_CODE is required.")
            continue

        if debit > 0 and credit > 0:
            errors.append(f"Row {row_number}: cannot have both debit and credit.")
            continue

        if debit <= 0 and credit <= 0:
            errors.append(f"Row {row_number}: debit or credit amount is required.")
            continue

        account = ChartOfAccount.objects.filter(
            company=company,
            code=account_code,
            is_active=True,
            is_group=False,
        ).first()

        if not account:
            errors.append(f"Row {row_number}: account code {account_code} not found.")
            continue

        group_key = f"{entry_date}|{reference_no}|{entry_description}"

        if group_key not in grouped:
            grouped[group_key] = {
                "entry_date": entry_date,
                "reference_no": reference_no,
                "description": entry_description,
                "lines": [],
            }

        grouped[group_key]["lines"].append({
            "account": account,
            "description": memo_detail or entry_description,
            "debit": debit,
            "credit": credit,
            "row_number": row_number,
        })

    for group_key, data in grouped.items():
        lines = data["lines"]
        total_debit = sum((line["debit"] for line in lines), Decimal("0.00"))
        total_credit = sum((line["credit"] for line in lines), Decimal("0.00"))

        if len(lines) < 2:
            row_numbers = ", ".join(str(line["row_number"]) for line in lines)
            errors.append(f"Rows {row_numbers}: journal must have at least 2 lines.")
            continue

        if total_debit != total_credit:
            row_numbers = ", ".join(str(line["row_number"]) for line in lines)
            errors.append(
                f"Rows {row_numbers}: debit and credit not equal. "
                f"Debit {total_debit}, Credit {total_credit}."
            )
            continue

        entry = JournalEntry.objects.create(
            company=company,
            entry_date=data["entry_date"],
            reference_no=data["reference_no"],
            description=data["description"],
            status=get_posted_status(),
            created_by=user,
        )

        for line in lines:
            JournalEntryLine.objects.create(
                journal_entry=entry,
                account=line["account"],
                description=line["description"],
                debit=line["debit"],
                credit=line["credit"],
            )
            success += 1

    return success, len(rows), errors


def import_trial_balance(company, user, excel_file):
    return import_journal_style_opening(
        company=company,
        user=user,
        excel_file=excel_file,
        default_description="Trial Balance Opening Balance",
    )


def import_journal_opening(company, user, excel_file):
    return import_journal_style_opening(
        company=company,
        user=user,
        excel_file=excel_file,
        default_description="Journal Opening Balance",
    )


@transaction.atomic
def import_outstanding_ap(company, user, excel_file):
    """
    Supports Batch_Outstanding_AP.xlsx sample.

    Normal PURCHASE:
        Dr Opening Balance Equity
        Cr Accounts Payable

    PURCHASE_RETURN:
        Dr Accounts Payable
        Cr Opening Balance Equity
    """
    rows, errors = read_excel_rows(excel_file)
    success = 0

    if errors:
        return success, len(rows), errors

    ap_account = get_account_by_code_or_name(
        company,
        codes=["290100", "200000"],
        names=["Accounts Payable", "Account Payable"],
    )
    opening_account = get_account_by_code_or_name(
        company,
        codes=["399999"],
        names=["Opening Balance", "Retained Earnings"],
    )

    if not ap_account:
        errors.append("Accounts Payable account not found. Please create/match account code 290100 or 200000.")

    if not opening_account:
        errors.append("Opening Balance Equity account not found. Please create/match account code 399999.")

    if errors:
        return success, len(rows), errors

    for row in rows:
        row_number = row["_ROW"]

        tran_type = clean_text(get_value(row, "TRAN_TYPE")).upper()
        vendor = clean_text(get_value(row, "VENDOR", "VENDOR_NAME"))
        bill_no = clean_text(get_value(row, "BILL_NO", "NUMBER", "INVOICE_NO"))
        memo = clean_text(get_value(row, "MEMO")) or "Opening AP Balance"
        entry_date = parse_excel_date(get_value(row, "DATE"))
        amount = to_decimal(get_value(row, "AMOUNT", "GROSS_NET", "COST", "AMOUNT_EXPENSE"))

        if not vendor and amount == 0:
            continue

        if not vendor:
            errors.append(f"Row {row_number}: VENDOR is required.")
            continue

        if amount <= 0:
            errors.append(f"Row {row_number}: AMOUNT must be more than zero.")
            continue

        reference_no = bill_no or f"AP-{row_number}"
        description = f"{memo} - {vendor}"

        if tran_type == "PURCHASE_RETURN":
            debit_account = ap_account
            credit_account = opening_account
        else:
            debit_account = opening_account
            credit_account = ap_account

        create_simple_journal(
            company=company,
            user=user,
            entry_date=entry_date,
            reference_no=reference_no,
            description=description,
            debit_account=debit_account,
            credit_account=credit_account,
            amount=amount,
        )

        success += 1

    return success, len(rows), errors


@transaction.atomic
def import_outstanding_ar(company, user, excel_file):
    """
    Supports Batch_Outstanding_AR.xlsx sample.

    INVOICE:
        Dr Accounts Receivable
        Cr Opening Balance Equity
    """
    rows, errors = read_excel_rows(excel_file)
    success = 0

    if errors:
        return success, len(rows), errors

    ar_account = get_account_by_code_or_name(
        company,
        codes=["150100", "120000"],
        names=["Accounts Receivable", "Account Receivable"],
    )
    opening_account = get_account_by_code_or_name(
        company,
        codes=["399999"],
        names=["Opening Balance", "Retained Earnings"],
    )

    if not ar_account:
        errors.append("Accounts Receivable account not found. Please create/match account code 150100 or 120000.")

    if not opening_account:
        errors.append("Opening Balance Equity account not found. Please create/match account code 399999.")

    if errors:
        return success, len(rows), errors

    for row in rows:
        row_number = row["_ROW"]

        customer = clean_text(get_value(row, "CUSTOMER", "CUSTOMER_NAME"))
        number = clean_text(get_value(row, "NUMBER", "INVOICE_NO"))
        memo = clean_text(get_value(row, "MEMO")) or "Opening AR Balance"
        entry_date = parse_excel_date(get_value(row, "DATE"))
        amount = to_decimal(get_value(row, "AMOUNT", "GROSS_NET", "PRICE"))

        if not customer and amount == 0:
            continue

        if not customer:
            errors.append(f"Row {row_number}: CUSTOMER is required.")
            continue

        if amount <= 0:
            errors.append(f"Row {row_number}: AMOUNT must be more than zero.")
            continue

        reference_no = number or f"AR-{row_number}"
        description = f"{memo} - {customer}"

        create_simple_journal(
            company=company,
            user=user,
            entry_date=entry_date,
            reference_no=reference_no,
            description=description,
            debit_account=ar_account,
            credit_account=opening_account,
            amount=amount,
        )

        success += 1

    return success, len(rows), errors


# =========================================================
# ROUTER
# =========================================================

def process_import(company, user, import_type, excel_file):
    from .models import ImportHistory

    if import_type == ImportHistory.TYPE_STOCK_BALANCE:
        return import_stock_balance(company, user, excel_file)

    if import_type == ImportHistory.TYPE_ITEM:
        return import_items(company, user, excel_file)

    if import_type == ImportHistory.TYPE_VENDOR:
        return import_vendors(company, user, excel_file)

    if import_type == ImportHistory.TYPE_CUSTOMER:
        return import_customers(company, user, excel_file)

    if import_type == ImportHistory.TYPE_COA:
        return import_chart_of_accounts(company, user, excel_file)

    if import_type == ImportHistory.TYPE_BATCH_TRANSACTION:
        return import_batch_transaction(company, user, excel_file)

    return 0, 0, ["Unknown import type."]