from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q, Sum
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from accounting.models import JournalEntry, JournalEntryLine
from core.models import Company

from .forms import (
    CustomerForm,
    CustomerTransactionForm,
    CustomerTypeForm,
    PriceLevelForm,
    RegionForm,
    SalesDocumentForm,
    SalespersonForm,
)
from .models import (
    Customer,
    CustomerTransaction,
    CustomerType,
    PriceLevel,
    Region,
    SalesDocument,
    Salesperson,
)


def get_selected_company(request):
    company_id = request.session.get("selected_company_id")
    if not company_id:
        return None
    return Company.objects.filter(id=company_id, is_active=True).first()


def can_access_company(user, company):
    if not company:
        return False

    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)

    if profile and getattr(profile, "user_type", None) == "client":
        return profile.company_id == company.id

    if hasattr(company, "assigned_staff"):
        return company.assigned_staff.filter(id=user.id).exists()

    return False


def require_company_access(request):
    company = get_selected_company(request)

    if not company:
        messages.warning(request, "Please select a company first.")
        return None, redirect("company_list")

    if not can_access_company(request.user, company):
        messages.error(request, "You do not have permission to access this company.")
        return None, redirect("company_list")

    return company, None


def get_posted_status():
    return getattr(JournalEntry, "STATUS_POSTED", "posted")


def create_customer_journal(customer_transaction, user):
    if customer_transaction.status != CustomerTransaction.STATUS_POSTED:
        return None

    if not customer_transaction.debit_account or not customer_transaction.credit_account:
        return None

    amount = customer_transaction.amount

    if amount <= 0:
        return None

    with transaction.atomic():
        old_entry = customer_transaction.journal_entry

        if old_entry:
            old_entry.delete()

        entry = JournalEntry.objects.create(
            company=customer_transaction.company,
            entry_date=customer_transaction.transaction_date,
            reference_no=customer_transaction.number or customer_transaction.so_number,
            description=f"{customer_transaction.get_transaction_type_display()} - {customer_transaction.customer.name}",
            status=get_posted_status(),
            created_by=user,
        )

        JournalEntryLine.objects.create(
            journal_entry=entry,
            account=customer_transaction.debit_account,
            description=customer_transaction.memo or customer_transaction.get_transaction_type_display(),
            debit=amount,
            credit=Decimal("0.00"),
        )

        JournalEntryLine.objects.create(
            journal_entry=entry,
            account=customer_transaction.credit_account,
            description=customer_transaction.memo or customer_transaction.get_transaction_type_display(),
            debit=Decimal("0.00"),
            credit=amount,
        )

        customer_transaction.journal_entry = entry
        customer_transaction.save(update_fields=["journal_entry"])

        return entry


@login_required
def customer_center(request):
    company, response = require_company_access(request)
    if response:
        return response

    query = (request.GET.get("q") or "").strip()

    customers = Customer.objects.filter(company=company)

    if query:
        customers = customers.filter(
            Q(name__icontains=query)
            | Q(code__icontains=query)
            | Q(phone__icontains=query)
            | Q(email__icontains=query)
            | Q(telegram__icontains=query)
        )

    customers = customers.select_related(
        "customer_type",
        "salesperson",
        "price_level",
        "region",
    ).order_by("name")

    posted_transactions = CustomerTransaction.objects.filter(
        company=company,
        status=CustomerTransaction.STATUS_POSTED,
    )

    total_invoice = posted_transactions.filter(
        transaction_type=CustomerTransaction.TYPE_INVOICE,
    ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

    total_payment = posted_transactions.filter(
        transaction_type=CustomerTransaction.TYPE_RECEIVE_PAYMENT,
    ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

    # Include the new customer-required Invoice / Receipt workflow while preserving legacy history.
    try:
        total_invoice += SalesInvoice.objects.filter(company=company, document_type=SalesInvoice.TYPE_INVOICE, status=SalesInvoice.STATUS_POSTED).aggregate(total=Sum("total_amount"))["total"] or Decimal("0.00")
        total_payment += CustomerReceipt.objects.filter(company=company, status=CustomerReceipt.STATUS_POSTED).aggregate(total=Sum("total_amount"))["total"] or Decimal("0.00")
    except Exception:
        # Before the new migration is applied, keep the legacy dashboard usable.
        pass
    ar_balance = total_invoice - total_payment

    return render(request, "customers/customer_center.html", {
        "company": company,
        "customers": customers,
        "query": query,
        "total_customers": customers.count(),
        "total_invoice": total_invoice,
        "total_payment": total_payment,
        "ar_balance": ar_balance,
    })


@login_required
def customer_create(request):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method == "POST":
        form = CustomerForm(request.POST, company=company)

        if form.is_valid():
            customer = form.save(commit=False)
            customer.company = company
            customer.created_by = request.user
            customer.save()

            messages.success(request, "Customer created successfully.")
            return redirect("customer_center")
    else:
        form = CustomerForm(company=company, initial={"is_active": True})

    return render(request, "customers/customer_form.html", {
        "company": company,
        "form": form,
        "page_title": "Create Customer",
        "button_text": "Create Customer",
    })


@login_required
def customer_edit(request, customer_id):
    company, response = require_company_access(request)
    if response:
        return response

    customer = get_object_or_404(Customer, id=customer_id, company=company)

    if request.method == "POST":
        form = CustomerForm(request.POST, instance=customer, company=company)

        if form.is_valid():
            form.save()
            messages.success(request, "Customer updated successfully.")
            return redirect("customer_center")
    else:
        form = CustomerForm(instance=customer, company=company)

    return render(request, "customers/customer_form.html", {
        "company": company,
        "form": form,
        "customer": customer,
        "page_title": "Edit Customer",
        "button_text": "Save Changes",
    })


@login_required
def customer_transaction_list(request):
    company, response = require_company_access(request)
    if response:
        return response

    query = (request.GET.get("q") or "").strip()
    tran_type = (request.GET.get("type") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    if not date_from and not date_to:
        today = timezone.localdate()
        date_from = today.replace(day=1).strftime("%Y-%m-%d")
        date_to = today.strftime("%Y-%m-%d")

    transactions_qs = CustomerTransaction.objects.filter(company=company)

    if query:
        transactions_qs = transactions_qs.filter(
            Q(customer__name__icontains=query)
            | Q(number__icontains=query)
            | Q(so_number__icontains=query)
            | Q(memo__icontains=query)
        )

    if tran_type:
        transactions_qs = transactions_qs.filter(transaction_type=tran_type)

    if date_from:
        transactions_qs = transactions_qs.filter(transaction_date__gte=date_from)

    if date_to:
        transactions_qs = transactions_qs.filter(transaction_date__lte=date_to)

    transactions_qs = transactions_qs.select_related(
        "customer",
        "debit_account",
        "credit_account",
        "journal_entry",
    ).order_by("-transaction_date", "-id")

    total_amount = transactions_qs.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

    return render(request, "customers/customer_transaction_list.html", {
        "company": company,
        "transactions": transactions_qs,
        "query": query,
        "tran_type": tran_type,
        "date_from": date_from,
        "date_to": date_to,
        "total_amount": total_amount,
        "type_choices": CustomerTransaction.TYPE_CHOICES,
    })


@login_required
def customer_transaction_create(request, transaction_type=None):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method == "POST":
        form = CustomerTransactionForm(
            request.POST,
            company=company,
            transaction_type=transaction_type,
        )

        if form.is_valid():
            with transaction.atomic():
                customer_transaction = form.save(commit=False)
                customer_transaction.company = company
                customer_transaction.created_by = request.user

                if transaction_type:
                    customer_transaction.transaction_type = transaction_type

                customer_transaction.save()
                create_customer_journal(customer_transaction, request.user)

            messages.success(request, "Customer transaction saved and journal generated.")
            return redirect("customer_transaction_list")
    else:
        form = CustomerTransactionForm(
            company=company,
            transaction_type=transaction_type,
            initial={
                "transaction_date": timezone.localdate(),
                "status": CustomerTransaction.STATUS_POSTED,
                "transaction_type": transaction_type or CustomerTransaction.TYPE_INVOICE,
                "currency": "USD",
                "exchange_rate": 1,
            },
        )

    title_map = {
        CustomerTransaction.TYPE_INVOICE: "Create Customer Invoice",
        CustomerTransaction.TYPE_RECEIVE_PAYMENT: "Create Receive Payment",
        CustomerTransaction.TYPE_CREDIT_NOTE: "Create Credit Note",
        CustomerTransaction.TYPE_ADJUSTMENT: "Create Customer Adjustment",
    }

    return render(request, "customers/customer_transaction_form.html", {
        "company": company,
        "form": form,
        "page_title": title_map.get(transaction_type, "Create Customer Transaction"),
        "button_text": "Save & Generate Journal",
    })


@login_required
def customer_transaction_detail(request, transaction_id):
    company, response = require_company_access(request)
    if response:
        return response

    customer_transaction = get_object_or_404(
        CustomerTransaction.objects.select_related(
            "customer",
            "debit_account",
            "credit_account",
            "journal_entry",
        ),
        id=transaction_id,
        company=company,
    )

    return render(request, "customers/customer_transaction_detail.html", {
        "company": company,
        "transaction": customer_transaction,
    })


@login_required
def sales_document_list(request, document_type):
    company, response = require_company_access(request)
    if response:
        return response

    query = (request.GET.get("q") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()
    customer_id = (request.GET.get("customer") or "").strip()

    if not date_from and not date_to:
        today = timezone.localdate()
        date_from = today.replace(day=1).strftime("%Y-%m-%d")
        date_to = today.strftime("%Y-%m-%d")

    docs = SalesDocument.objects.filter(company=company, document_type=document_type)

    if query:
        docs = docs.filter(
            Q(number__icontains=query)
            | Q(customer__name__icontains=query)
            | Q(memo__icontains=query)
        )

    if customer_id:
        docs = docs.filter(customer_id=customer_id)

    if date_from:
        docs = docs.filter(document_date__gte=date_from)

    if date_to:
        docs = docs.filter(document_date__lte=date_to)

    docs = docs.select_related("customer", "salesperson").order_by("-document_date", "-id")

    title = "Quotation" if document_type == SalesDocument.TYPE_QUOTATION else "Sale Order"

    return render(request, "customers/sales_document_list.html", {
        "company": company,
        "docs": docs,
        "document_type": document_type,
        "title": title,
        "query": query,
        "date_from": date_from,
        "date_to": date_to,
        "customer_id": customer_id,
        "customers": Customer.objects.filter(company=company, is_active=True).order_by("name"),
    })


@login_required
def sales_document_create(request, document_type):
    company, response = require_company_access(request)
    if response:
        return response

    sales_doc = SalesDocument(
        company=company,
        document_type=document_type,
        created_by=request.user,
    )

    if request.method == "POST":
        form = SalesDocumentForm(request.POST, instance=sales_doc, company=company)

        if form.is_valid():
            doc = form.save(commit=False)
            doc.company = company
            doc.document_type = document_type
            doc.created_by = request.user
            doc.save()

            messages.success(request, "Sales document saved successfully.")
            return redirect("sales_document_list", document_type=document_type)
    else:
        form = SalesDocumentForm(
            instance=sales_doc,
            company=company,
            initial={
                "document_date": timezone.localdate(),
                "status": SalesDocument.STATUS_OPEN,
                "currency": "USD",
            },
        )

    title = "Create Quotation" if document_type == SalesDocument.TYPE_QUOTATION else "Create Sale Order"

    return render(request, "customers/sales_document_form.html", {
        "company": company,
        "form": form,
        "document_type": document_type,
        "page_title": title,
        "button_text": "Save",
    })


def master_list_view(request, model, template, title):
    company, response = require_company_access(request)
    if response:
        return response

    query = (request.GET.get("q") or "").strip()
    rows = model.objects.filter(company=company)

    if query:
        if model == CustomerType:
            rows = rows.filter(
                Q(name__icontains=query)
                | Q(memo__icontains=query)
            )
        elif model == Salesperson:
            rows = rows.filter(
                Q(name__icontains=query)
                | Q(code__icontains=query)
                | Q(local_name__icontains=query)
                | Q(phone__icontains=query)
                | Q(email__icontains=query)
                | Q(memo__icontains=query)
            )
        elif model == PriceLevel:
            rows = rows.filter(
                Q(name__icontains=query)
                | Q(round_type__icontains=query)
                | Q(discount_method__icontains=query)
                | Q(memo__icontains=query)
            )
        else:
            rows = rows.filter(
                Q(name__icontains=query)
                | Q(code__icontains=query)
                | Q(local_name__icontains=query)
                | Q(memo__icontains=query)
            )

    return render(request, template, {
        "company": company,
        "rows": rows.order_by("name"),
        "query": query,
        "title": title,
    })


def master_create_view(request, form_class, template, title, redirect_name):
    company, response = require_company_access(request)
    if response:
        return response

    if request.method == "POST":
        form = form_class(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = company
            obj.save()

            messages.success(request, f"{title} saved successfully.")
            return redirect(redirect_name)
    else:
        form = form_class(initial={"is_active": True})

    return render(request, template, {
        "company": company,
        "form": form,
        "page_title": title,
        "button_text": "Save",
    })


@login_required
def customer_type_list(request):
    return master_list_view(request, CustomerType, "customers/customer_type_list.html", "Customer Type")


@login_required
def salesperson_list(request):
    return master_list_view(request, Salesperson, "customers/salesperson_list.html", "Salesperson")


@login_required
def price_level_list(request):
    return master_list_view(request, PriceLevel, "customers/price_level_list.html", "Price Level")


@login_required
def region_list(request):
    return master_list_view(request, Region, "customers/region_list.html", "Region")


@login_required
def customer_type_create(request):
    return master_create_view(request, CustomerTypeForm, "customers/master_form.html", "Create Customer Type", "customer_type_list")


@login_required
def salesperson_create(request):
    return master_create_view(request, SalespersonForm, "customers/master_form.html", "Create Salesperson", "salesperson_list")


@login_required
def price_level_create(request):
    return master_create_view(request, PriceLevelForm, "customers/master_form.html", "Create Price Level", "price_level_list")


@login_required
def region_create(request):
    return master_create_view(request, RegionForm, "customers/master_form.html", "Create Region", "region_list")


@login_required
def quotation_list(request):
    return sales_document_list(request, SalesDocument.TYPE_QUOTATION)


@login_required
def sale_order_list(request):
    return sales_document_list(request, SalesDocument.TYPE_SALE_ORDER)


@login_required
def quotation_create(request):
    return sales_document_create(request, SalesDocument.TYPE_QUOTATION)


@login_required
def sale_order_create(request):
    return sales_document_create(request, SalesDocument.TYPE_SALE_ORDER)


@login_required
def invoice_create(request):
    return customer_transaction_create(request, CustomerTransaction.TYPE_INVOICE)


@login_required
def receive_payment_create(request):
    return customer_transaction_create(request, CustomerTransaction.TYPE_RECEIVE_PAYMENT)
# =========================================================
# CUSTOMER REQUIREMENT IMPLEMENTATION
# =========================================================
from django.http import HttpResponse
from django.forms import modelformset_factory
from openpyxl import Workbook, load_workbook
from stock.models import Item
from .forms import SalesInvoiceForm, SalesInvoiceLineFormSet, CustomerReceiptForm, CustomerReceiptAllocationFormSet, CustomerReceiptOtherChargeFormSet
from .models import SalesInvoice, SalesInvoiceLine, CustomerReceipt, CustomerReceiptAllocation, CustomerReceiptOtherCharge


def _formset_with_company(formset_class, data, instance, company, extra_kwargs=None, prefix=None):
    extra_kwargs = extra_kwargs or {}
    kwargs = {"instance": instance, "prefix": prefix, "form_kwargs": {"company": company, **extra_kwargs}}
    if data is not None:
        kwargs["data"] = data
    return formset_class(**kwargs)


def create_sales_invoice_journal(invoice, user):
    if invoice.status != SalesInvoice.STATUS_POSTED or invoice.total_amount <= 0:
        return None
    with transaction.atomic():
        if invoice.journal_entry_id:
            invoice.journal_entry.delete()
        entry = JournalEntry.objects.create(company=invoice.company, entry_date=invoice.invoice_date,
            reference_no=invoice.number, description=f"{invoice.get_document_type_display()} - {invoice.customer.name}", status=get_posted_status(), created_by=user)
        debit_account = invoice.deposit_account if invoice.document_type == SalesInvoice.TYPE_SALE_RECEIPT else invoice.accounts_receivable_account
        JournalEntryLine.objects.create(journal_entry=entry, account=debit_account, description=invoice.memo or invoice.get_document_type_display(), debit=invoice.total_amount, credit=Decimal("0.00"))
        for line in invoice.lines.select_related("revenue_account"):
            net = (line.line_amount or Decimal("0")) + (line.tax_amount or Decimal("0")) - (line.discount_amount or Decimal("0"))
            if net > 0:
                JournalEntryLine.objects.create(journal_entry=entry, account=line.revenue_account, description=line.description or line.item.name, debit=Decimal("0.00"), credit=net)
        invoice.journal_entry = entry; invoice.save(update_fields=["journal_entry"])
    return entry


def create_customer_receipt_journal(receipt, user):
    if receipt.status != CustomerReceipt.STATUS_POSTED or receipt.total_amount <= 0:
        return None
    with transaction.atomic():
        if receipt.journal_entry_id: receipt.journal_entry.delete()
        entry=JournalEntry.objects.create(company=receipt.company,entry_date=receipt.receipt_date,reference_no=receipt.number,description=f"Receipt / Collection - {receipt.customer.name}",status=get_posted_status(),created_by=user)
        JournalEntryLine.objects.create(journal_entry=entry,account=receipt.deposit_account,description=receipt.memo or "Customer receipt",debit=receipt.total_amount,credit=Decimal("0.00"))
        ar_account=None
        for alloc in receipt.allocations.select_related("invoice__accounts_receivable_account"):
            ar_account=alloc.invoice.accounts_receivable_account
            credit=(alloc.amount or Decimal("0"))+(alloc.discount or Decimal("0"))
            if credit>0: JournalEntryLine.objects.create(journal_entry=entry,account=ar_account,description=f"Apply {alloc.invoice.number}",debit=Decimal("0.00"),credit=credit)
        for ch in receipt.other_charges.select_related("account"):
            if ch.amount>0: JournalEntryLine.objects.create(journal_entry=entry,account=ch.account,description=ch.memo,debit=Decimal("0.00"),credit=ch.amount)
        receipt.journal_entry=entry;receipt.save(update_fields=["journal_entry"])
    return entry


@login_required
def customer_export_excel(request):
    company, response = require_company_access(request)
    if response:
        return response

    # Use the same search as Customer Center so the Excel file matches
    # what the user is currently viewing.
    query = (request.GET.get("q") or "").strip()

    customers = Customer.objects.filter(company=company)

    if query:
        customers = customers.filter(
            Q(name__icontains=query)
            | Q(code__icontains=query)
            | Q(phone__icontains=query)
            | Q(email__icontains=query)
            | Q(telegram__icontains=query)
        )

    customers = customers.select_related(
        "salesperson",
        "region",
    ).order_by("name")

    wb = Workbook()
    ws = wb.active
    ws.title = "Customer List"

    ws.append([
        "Customer",
        "Code",
        "Phone",
        "Email",
        "Salesperson",
        "Region",
        "A/R Balance",
    ])

    for customer in customers:
        ws.append([
            customer.name or "",
            customer.code or "",
            customer.phone or "",
            customer.email or "",
            customer.salesperson.name if customer.salesperson else "",
            customer.region.name if customer.region else "",
            float(customer.ar_balance or 0),
        ])

    # Make the downloaded sheet easier to read.
    widths = {
        "A": 30,
        "B": 16,
        "C": 18,
        "D": 30,
        "E": 22,
        "F": 22,
        "G": 16,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    for cell in ws["G"][1:]:
        cell.number_format = '#,##0.00'

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="customer_list.xlsx"'
    wb.save(response)
    return response


@login_required
def customer_import_excel(request):
    company,response=require_company_access(request)
    if response:return response
    if request.method=="POST" and request.FILES.get("file"):
        try:
            wb=load_workbook(request.FILES["file"],data_only=True);ws=wb.active;created=updated=0
            headers=[str(x.value or "").strip().lower() for x in ws[1]]
            def idx(*names):
                for n in names:
                    if n in headers:return headers.index(n)
                return None
            name_i=idx("customer name","name");code_i=idx("customer id","code")
            if name_i is None: raise ValueError("Excel needs a Customer Name column.")
            for row in ws.iter_rows(min_row=2,values_only=True):
                name=str(row[name_i] or "").strip()
                if not name:continue
                code=str(row[code_i] or "").strip() if code_i is not None else ""
                defaults={"code":code}
                for field,names in [("phone",("phone",)),("email",("email",)),("telegram",("telegram",)),("address",("address",)),("local_name",("local name",))]:
                    i=idx(*names)
                    if i is not None:defaults[field]=str(row[i] or "").strip()
                obj,was_created=Customer.objects.update_or_create(company=company,name=name,defaults=defaults)
                created+=int(was_created);updated+=int(not was_created)
            messages.success(request,f"Customer import complete: {created} created, {updated} updated.");return redirect("customer_center")
        except Exception as exc:messages.error(request,f"Import failed: {exc}")
    return render(request,"customers/customer_import.html",{"company":company})


@login_required
def sales_invoice_list(request, document_type=SalesInvoice.TYPE_INVOICE):
    company,response=require_company_access(request)
    if response:return response
    qs=SalesInvoice.objects.filter(company=company,document_type=document_type).select_related("customer")
    q=(request.GET.get("q") or "").strip()
    if q:qs=qs.filter(Q(number__icontains=q)|Q(customer__name__icontains=q)|Q(customer__code__icontains=q)|Q(po_number__icontains=q))
    return render(request,"customers/invoice_list.html",{"company":company,"invoices":qs,"document_type":document_type,"query":q,"page_title":"Sale Receipt" if document_type==SalesInvoice.TYPE_SALE_RECEIPT else "Sale Invoices"})


def _invoice_formsets(request,company,invoice):
    return _formset_with_company(SalesInvoiceLineFormSet,request.POST if request.method=="POST" else None,invoice,company,prefix="lines")

@login_required
def sales_invoice_create(request, document_type=SalesInvoice.TYPE_INVOICE):
    company,response=require_company_access(request)
    if response:return response
    invoice=SalesInvoice(company=company,document_type=document_type,created_by=request.user)
    form=SalesInvoiceForm(request.POST or None,instance=invoice,company=company,document_type=document_type)
    lines=_invoice_formsets(request,company,invoice)
    if request.method=="POST" and form.is_valid() and lines.is_valid():
        try:
            with transaction.atomic():
                invoice=form.save(commit=False);invoice.company=company;invoice.document_type=document_type;invoice.created_by=request.user;invoice.status=SalesInvoice.STATUS_POSTED;invoice.save()
                lines.instance=invoice;lines.save();invoice.recalculate_totals();create_sales_invoice_journal(invoice,request.user)
            messages.success(request,f"{invoice.get_document_type_display()} saved successfully.")
            if request.POST.get("save_action")=="save_new":return redirect("sale_receipt_create" if document_type==SalesInvoice.TYPE_SALE_RECEIPT else "customer_invoice_create")
            return redirect("sale_receipt_list" if document_type==SalesInvoice.TYPE_SALE_RECEIPT else "customer_invoice_list")
        except Exception as exc:messages.error(request,f"Could not save: {exc}")
    return render(request,"customers/invoice_form.html",{"company":company,"form":form,"line_formset":lines,"document_type":document_type,"page_title":"Sale Receipt" if document_type==SalesInvoice.TYPE_SALE_RECEIPT else "Sale Invoice"})


@login_required
def customer_receipt_create(request):
    company,response=require_company_access(request)
    if response:return response
    receipt=CustomerReceipt(company=company,created_by=request.user)
    customer_id=(request.POST.get("customer") if request.method=="POST" else request.GET.get("customer")) or None
    form=CustomerReceiptForm(request.POST or None,instance=receipt,company=company)
    allocations=_formset_with_company(CustomerReceiptAllocationFormSet,request.POST if request.method=="POST" else None,receipt,company,{"customer_id":customer_id},"alloc")
    charges=_formset_with_company(CustomerReceiptOtherChargeFormSet,request.POST if request.method=="POST" else None,receipt,company,prefix="charge")
    if request.method=="POST" and form.is_valid() and allocations.is_valid() and charges.is_valid():
        try:
            with transaction.atomic():
                receipt=form.save(commit=False);receipt.company=company;receipt.created_by=request.user;receipt.status=CustomerReceipt.STATUS_POSTED;receipt.save()
                allocations.instance=receipt;allocations.save();charges.instance=receipt;charges.save();receipt.recalculate_total();create_customer_receipt_journal(receipt,request.user)
            messages.success(request,"Receipt / Collection saved.")
            if request.POST.get("save_action")=="save_new":return redirect("receive_payment_create")
            return redirect("customer_center")
        except Exception as exc:messages.error(request,f"Could not save receipt: {exc}")
    open_invoices=SalesInvoice.objects.filter(company=company,document_type=SalesInvoice.TYPE_INVOICE,status=SalesInvoice.STATUS_POSTED)
    if customer_id:open_invoices=open_invoices.filter(customer_id=customer_id)
    return render(request,"customers/receipt_form.html",{"company":company,"form":form,"allocation_formset":allocations,"charge_formset":charges,"open_invoices":open_invoices})
